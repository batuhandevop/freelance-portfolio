"""
Book Scraper - Web Scraping Portfolio Example
==============================================
Scrapes book data from https://books.toscrape.com and exports to CSV and Excel.

Requirements:
    pip install requests beautifulsoup4 openpyxl

Usage:
    python book_scraper.py
"""

import csv
import time
import os
from dataclasses import dataclass, fields, asdict

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_URL = "https://books.toscrape.com"
CATALOGUE_URL = f"{BASE_URL}/catalogue"
MAX_PAGES = 5  # Set to None to scrape all pages
REQUEST_DELAY = 1.0  # Seconds between requests (be polite)
MAX_RETRIES = 3
RETRY_BACKOFF = 2.0  # Multiplier for exponential backoff
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

HEADERS = {
    "User-Agent": "BookScraper/1.0 (portfolio-demo; educational-purposes)"
}

# Rating words on the site map to numeric values
RATING_MAP = {
    "One": 1,
    "Two": 2,
    "Three": 3,
    "Four": 4,
    "Five": 5,
}


# ---------------------------------------------------------------------------
# Data Model
# ---------------------------------------------------------------------------

@dataclass
class Book:
    title: str
    price: float
    rating: int
    availability: str
    category: str


# ---------------------------------------------------------------------------
# HTTP Helpers
# ---------------------------------------------------------------------------

def fetch_page(url: str) -> BeautifulSoup | None:
    """Fetch a URL with retry logic and return parsed HTML."""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = requests.get(url, headers=HEADERS, timeout=15)
            response.raise_for_status()
            return BeautifulSoup(response.text, "html.parser")
        except requests.RequestException as exc:
            wait = RETRY_BACKOFF ** attempt
            print(f"  [!] Request failed (attempt {attempt}/{MAX_RETRIES}): {exc}")
            if attempt < MAX_RETRIES:
                print(f"      Retrying in {wait:.0f}s ...")
                time.sleep(wait)
    print(f"  [X] Giving up on {url}")
    return None


# ---------------------------------------------------------------------------
# Scraping Logic
# ---------------------------------------------------------------------------

def get_book_category(detail_url: str) -> str:
    """Visit a book's detail page to extract its category."""
    soup = fetch_page(detail_url)
    if soup is None:
        return "Unknown"

    # Category sits in the breadcrumb: Home > Books > <Category> > <Title>
    breadcrumb_links = soup.select("ul.breadcrumb li a")
    if len(breadcrumb_links) >= 3:
        return breadcrumb_links[2].get_text(strip=True)
    return "Unknown"


def parse_book(article, fetch_categories: bool = True) -> Book:
    """Parse a single <article class='product_pod'> element into a Book."""

    # --- Title ---
    title_tag = article.select_one("h3 a")
    title = title_tag["title"] if title_tag else "N/A"

    # --- Price ---
    price_text = article.select_one(".price_color").get_text(strip=True)
    price = float(price_text.replace("£", "").replace("Â", ""))

    # --- Rating ---
    star_tag = article.select_one("p.star-rating")
    rating_class = star_tag["class"][1] if star_tag else "Zero"
    rating = RATING_MAP.get(rating_class, 0)

    # --- Availability ---
    avail_tag = article.select_one(".availability")
    availability = avail_tag.get_text(strip=True) if avail_tag else "Unknown"

    # --- Category (requires visiting the detail page) ---
    category = "Unknown"
    if fetch_categories and title_tag:
        relative_href = title_tag["href"]
        # Links on listing pages look like: ../catalogue/<slug>/index.html
        detail_url = f"{CATALOGUE_URL}/{relative_href.replace('../', '')}"
        category = get_book_category(detail_url)
        time.sleep(REQUEST_DELAY)

    return Book(
        title=title,
        price=price,
        rating=rating,
        availability=availability,
        category=category,
    )


def scrape_books(max_pages: int | None = MAX_PAGES,
                 fetch_categories: bool = True) -> list[Book]:
    """Scrape book listings across multiple pages."""
    books: list[Book] = []
    page = 1

    while True:
        if max_pages and page > max_pages:
            break

        url = (
            f"{CATALOGUE_URL}/page-{page}.html" if page > 1
            else f"{BASE_URL}/index.html"
        )
        print(f"[*] Scraping page {page}: {url}")
        soup = fetch_page(url)

        if soup is None:
            print("    Could not load page, stopping pagination.")
            break

        articles = soup.select("article.product_pod")
        if not articles:
            print("    No books found on this page, stopping.")
            break

        for i, article in enumerate(articles, start=1):
            book = parse_book(article, fetch_categories=fetch_categories)
            books.append(book)
            print(f"    [{len(books):>4}] {book.title[:50]:<50} "
                  f"£{book.price:>5.2f}  {'*' * book.rating}")

        # Check for a "next" button
        next_btn = soup.select_one("li.next a")
        if next_btn is None:
            print("[*] Reached the last page.")
            break

        page += 1
        time.sleep(REQUEST_DELAY)

    return books


# ---------------------------------------------------------------------------
# Export Functions
# ---------------------------------------------------------------------------

def save_to_csv(books: list[Book], filename: str = "books.csv") -> str:
    """Save the book list to a CSV file."""
    filepath = os.path.join(OUTPUT_DIR, filename)
    field_names = [f.name for f in fields(Book)]

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=field_names)
        writer.writeheader()
        for book in books:
            writer.writerow(asdict(book))

    return filepath


def save_to_excel(books: list[Book], filename: str = "books.xlsx") -> str:
    """Save the book list to a formatted Excel file."""
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb = Workbook()
    ws = wb.active
    ws.title = "Books"

    # -- Header row styling --
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496",
                              fill_type="solid")
    header_align = Alignment(horizontal="center")

    field_names = [f.name for f in fields(Book)]
    headers = [name.replace("_", " ").title() for name in field_names]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # -- Data rows --
    for row_idx, book in enumerate(books, start=2):
        values = list(asdict(book).values())
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # -- Auto-fit column widths --
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row_idx in range(2, len(books) + 2):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, len(cell_val))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
            min(max_len + 3, 60)
        )

    # -- Freeze the header row --
    ws.freeze_panes = "A2"

    wb.save(filepath)
    return filepath


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 65)
    print("  Book Scraper - https://books.toscrape.com")
    print("=" * 65)

    start = time.time()
    books = scrape_books()
    elapsed = time.time() - start

    if not books:
        print("\nNo books were scraped. Check your connection and try again.")
        return

    print(f"\n[+] Scraped {len(books)} books in {elapsed:.1f}s")

    csv_path = save_to_csv(books)
    print(f"[+] CSV  saved to: {csv_path}")

    xlsx_path = save_to_excel(books)
    print(f"[+] Excel saved to: {xlsx_path}")

    # Quick summary
    avg_price = sum(b.price for b in books) / len(books)
    avg_rating = sum(b.rating for b in books) / len(books)
    categories = {b.category for b in books}
    print(f"\n--- Summary ---")
    print(f"  Total books : {len(books)}")
    print(f"  Avg price   : £{avg_price:.2f}")
    print(f"  Avg rating  : {avg_rating:.1f} / 5")
    print(f"  Categories  : {len(categories)}")
    print()


if __name__ == "__main__":
    main()
