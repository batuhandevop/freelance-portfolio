"""
Excel Report Generator
======================
Generates a professional multi-sheet monthly sales report using openpyxl.
Run: python report_generator.py
Output: monthly_report.xlsx
"""

import datetime
import random
from copy import copy

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

COMPANY_NAME = "Apex Analytics Ltd."
REPORT_TITLE = "Monthly Sales Report"
OUTPUT_FILE = "monthly_report.xlsx"

REGIONS = ["North America", "Europe", "Asia Pacific", "Latin America"]
PRODUCTS = ["Widget Pro", "Widget Lite", "Service Plan", "Accessory Pack", "Enterprise Suite"]
MONTHS = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]

# Company colour palette
COLOR_PRIMARY = "1B3A5C"       # dark navy
COLOR_SECONDARY = "2E86C1"    # medium blue
COLOR_ACCENT = "F39C12"       # amber
COLOR_LIGHT_BG = "EBF5FB"     # very light blue
COLOR_WHITE = "FFFFFF"
COLOR_DARK_TEXT = "1C1C1C"
COLOR_BORDER = "B0C4D8"

# Reusable style objects
THIN_BORDER = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color=COLOR_WHITE)
HEADER_FILL = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Calibri", size=10, color=COLOR_DARK_TEXT)
CURRENCY_FORMAT = '#,##0'
PERCENT_FORMAT = '0.0%'

# ---------------------------------------------------------------------------
# Data Generation
# ---------------------------------------------------------------------------


def generate_sales_data() -> list[dict]:
    """Create a realistic sample dataset of monthly sales by region and product."""
    random.seed(42)  # reproducible output
    base_prices = {
        "Widget Pro": 12000,
        "Widget Lite": 4500,
        "Service Plan": 8000,
        "Accessory Pack": 2000,
        "Enterprise Suite": 25000,
    }
    rows = []
    for month_idx, month in enumerate(MONTHS):
        for region in REGIONS:
            for product in PRODUCTS:
                # Add seasonal variation and regional weight
                seasonal = 1 + 0.15 * (month_idx / 11)  # slight upward trend
                regional_weight = {"North America": 1.3, "Europe": 1.1,
                                   "Asia Pacific": 0.9, "Latin America": 0.7}[region]
                noise = random.uniform(0.75, 1.30)
                revenue = int(base_prices[product] * seasonal * regional_weight * noise)
                units = max(1, revenue // (base_prices[product] // 10))
                rows.append({
                    "Month": month,
                    "MonthIndex": month_idx + 1,
                    "Region": region,
                    "Product": product,
                    "Units": units,
                    "Revenue": revenue,
                })
    return rows


# ---------------------------------------------------------------------------
# Styling Helpers
# ---------------------------------------------------------------------------


def style_header_row(ws, row: int, max_col: int):
    """Apply header styling to an entire row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def style_data_cell(cell, is_currency=False, is_alt_row=False):
    """Apply consistent data-cell styling."""
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if is_alt_row:
        cell.fill = PatternFill(start_color=COLOR_LIGHT_BG, end_color=COLOR_LIGHT_BG, fill_type="solid")
    if is_currency:
        cell.number_format = CURRENCY_FORMAT


def auto_width(ws, min_width=10, max_width=30):
    """Auto-fit column widths based on content length."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        length = max((len(str(c.value or "")) for c in col_cells), default=min_width)
        ws.column_dimensions[col_letter].width = min(max(length + 3, min_width), max_width)


# ---------------------------------------------------------------------------
# Sheet Builders
# ---------------------------------------------------------------------------


def build_cover_sheet(wb: Workbook):
    """Create a cover page with company branding."""
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_properties.tabColor = COLOR_PRIMARY

    # Remove gridlines for a clean look
    ws.sheet_view.showGridLines = False

    # Merge a wide area for the title block
    ws.merge_cells("B4:H6")
    title_cell = ws["B4"]
    title_cell.value = COMPANY_NAME
    title_cell.font = Font(name="Calibri", bold=True, size=28, color=COLOR_PRIMARY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B8:H9")
    subtitle = ws["B8"]
    subtitle.value = REPORT_TITLE
    subtitle.font = Font(name="Calibri", bold=True, size=20, color=COLOR_SECONDARY)
    subtitle.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B11:H12")
    date_cell = ws["B11"]
    date_cell.value = f"Generated: {datetime.date.today().strftime('%B %d, %Y')}"
    date_cell.font = Font(name="Calibri", size=14, color=COLOR_DARK_TEXT)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Decorative colour band
    band_fill = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
    for col in range(2, 9):
        for row in (3, 13):
            ws.cell(row=row, column=col).fill = band_fill

    # Set column widths
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14


def build_summary_sheet(wb: Workbook, data: list[dict]):
    """KPI summary sheet driven by formulas referencing the Detail sheet."""
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = COLOR_SECONDARY
    ws.sheet_view.showGridLines = False

    # --- Pre-compute values for KPI cards ---
    total_revenue = sum(r["Revenue"] for r in data)
    h1_revenue = sum(r["Revenue"] for r in data if r["MonthIndex"] <= 6)
    h2_revenue = sum(r["Revenue"] for r in data if r["MonthIndex"] > 6)
    growth = (h2_revenue - h1_revenue) / h1_revenue if h1_revenue else 0

    product_totals = {}
    for r in data:
        product_totals[r["Product"]] = product_totals.get(r["Product"], 0) + r["Revenue"]
    top_product = max(product_totals, key=product_totals.get)

    region_totals = {}
    for r in data:
        region_totals[r["Region"]] = region_totals.get(r["Region"], 0) + r["Revenue"]
    top_region = max(region_totals, key=region_totals.get)

    total_units = sum(r["Units"] for r in data)

    # --- Section title ---
    ws.merge_cells("B2:G2")
    hdr = ws["B2"]
    hdr.value = "Key Performance Indicators"
    hdr.font = Font(name="Calibri", bold=True, size=18, color=COLOR_PRIMARY)
    hdr.alignment = Alignment(horizontal="left")

    # Decorative line under title
    accent_fill = PatternFill(start_color=COLOR_ACCENT, end_color=COLOR_ACCENT, fill_type="solid")
    for c in range(2, 8):
        ws.cell(row=3, column=c).fill = accent_fill

    # --- KPI cards (label + value pairs) ---
    kpis = [
        ("Total Revenue", f"${total_revenue:,.0f}"),
        ("Total Units Sold", f"{total_units:,}"),
        ("H2 vs H1 Growth", f"{growth:.1%}"),
        ("Top Product", top_product),
        ("Top Region", top_region),
        ("Avg Revenue / Month", f"${total_revenue / 12:,.0f}"),
    ]

    card_fill = PatternFill(start_color=COLOR_LIGHT_BG, end_color=COLOR_LIGHT_BG, fill_type="solid")
    card_border = Border(
        left=Side(style="medium", color=COLOR_SECONDARY),
        right=Side(style="medium", color=COLOR_SECONDARY),
        top=Side(style="medium", color=COLOR_SECONDARY),
        bottom=Side(style="medium", color=COLOR_SECONDARY),
    )

    start_row = 5
    for i, (label, value) in enumerate(kpis):
        col_offset = 2 + (i % 3) * 2  # 3 cards per row
        row_offset = start_row + (i // 3) * 4

        # Label
        lbl_cell = ws.cell(row=row_offset, column=col_offset)
        lbl_cell.value = label
        lbl_cell.font = Font(name="Calibri", bold=True, size=10, color=COLOR_SECONDARY)
        lbl_cell.alignment = Alignment(horizontal="center")
        lbl_cell.fill = card_fill
        lbl_cell.border = card_border

        # Value
        val_cell = ws.cell(row=row_offset + 1, column=col_offset)
        val_cell.value = value
        val_cell.font = Font(name="Calibri", bold=True, size=14, color=COLOR_PRIMARY)
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.fill = card_fill
        val_cell.border = card_border

    # Column widths
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 22

    # --- Revenue by Product summary table ---
    table_start = 14
    ws.cell(row=table_start, column=2).value = "Revenue by Product"
    ws.cell(row=table_start, column=2).font = Font(name="Calibri", bold=True, size=14, color=COLOR_PRIMARY)

    headers = ["Product", "Revenue", "Share"]
    for ci, h in enumerate(headers, start=2):
        ws.cell(row=table_start + 1, column=ci).value = h
    style_header_row(ws, table_start + 1, 4)

    for ri, product in enumerate(PRODUCTS):
        row = table_start + 2 + ri
        alt = ri % 2 == 1
        rev = product_totals[product]

        c1 = ws.cell(row=row, column=2, value=product)
        style_data_cell(c1, is_alt_row=alt)

        c2 = ws.cell(row=row, column=3, value=rev)
        style_data_cell(c2, is_currency=True, is_alt_row=alt)

        c3 = ws.cell(row=row, column=4, value=rev / total_revenue)
        style_data_cell(c3, is_alt_row=alt)
        c3.number_format = PERCENT_FORMAT

    # --- Revenue by Region summary table ---
    table_start2 = table_start
    ws.cell(row=table_start2, column=6).value = "Revenue by Region"
    ws.cell(row=table_start2, column=6).font = Font(name="Calibri", bold=True, size=14, color=COLOR_PRIMARY)

    for ci, h in enumerate(["Region", "Revenue", "Share"], start=6):
        ws.cell(row=table_start2 + 1, column=ci).value = h
    style_header_row(ws, table_start2 + 1, 8)
    # only style cols 6-8
    for c in range(1, 6):
        cell = ws.cell(row=table_start2 + 1, column=c)
        # keep previous styling if any

    for ri, region in enumerate(REGIONS):
        row = table_start2 + 2 + ri
        alt = ri % 2 == 1
        rev = region_totals[region]

        c1 = ws.cell(row=row, column=6, value=region)
        style_data_cell(c1, is_alt_row=alt)

        c2 = ws.cell(row=row, column=7, value=rev)
        style_data_cell(c2, is_currency=True, is_alt_row=alt)

        c3 = ws.cell(row=row, column=8, value=rev / total_revenue)
        style_data_cell(c3, is_alt_row=alt)
        c3.number_format = PERCENT_FORMAT


def build_detail_sheet(wb: Workbook, data: list[dict]):
    """Full data table with filters and freeze panes."""
    ws = wb.create_sheet("Detail")
    ws.sheet_properties.tabColor = COLOR_ACCENT

    headers = ["Month", "Region", "Product", "Units", "Revenue"]
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=1, column=ci).value = h
    style_header_row(ws, 1, len(headers))

    for ri, row_data in enumerate(data):
        row_num = ri + 2
        alt = ri % 2 == 1
        values = [
            row_data["Month"],
            row_data["Region"],
            row_data["Product"],
            row_data["Units"],
            row_data["Revenue"],
        ]
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            is_currency = ci == 5
            style_data_cell(cell, is_currency=is_currency, is_alt_row=alt)

    # Auto-filter on the header row
    ws.auto_filter.ref = f"A1:E{len(data) + 1}"

    # Freeze the header row
    ws.freeze_panes = "A2"

    auto_width(ws)


def build_region_sheets(wb: Workbook, data: list[dict]):
    """One sheet per region with a filtered breakdown."""
    for region in REGIONS:
        # Sheet name limited to 31 chars
        sheet_name = region[:31]
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = COLOR_SECONDARY

        # Region title
        ws.merge_cells("A1:E1")
        title = ws["A1"]
        title.value = f"{region} -- Sales Detail"
        title.font = Font(name="Calibri", bold=True, size=14, color=COLOR_PRIMARY)
        title.alignment = Alignment(horizontal="left", vertical="center")

        headers = ["Month", "Product", "Units", "Revenue"]
        for ci, h in enumerate(headers, start=1):
            ws.cell(row=3, column=ci).value = h
        style_header_row(ws, 3, len(headers))

        region_rows = [r for r in data if r["Region"] == region]
        for ri, row_data in enumerate(region_rows):
            row_num = ri + 4
            alt = ri % 2 == 1
            values = [row_data["Month"], row_data["Product"],
                      row_data["Units"], row_data["Revenue"]]
            for ci, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=ci, value=val)
                style_data_cell(cell, is_currency=(ci == 4), is_alt_row=alt)

        # Totals row
        total_row = len(region_rows) + 4
        ws.cell(row=total_row, column=1).value = "TOTAL"
        ws.cell(row=total_row, column=1).font = Font(name="Calibri", bold=True, size=11)
        total_rev = sum(r["Revenue"] for r in region_rows)
        total_units = sum(r["Units"] for r in region_rows)
        ws.cell(row=total_row, column=3, value=total_units).font = Font(bold=True)
        rev_cell = ws.cell(row=total_row, column=4, value=total_rev)
        rev_cell.font = Font(name="Calibri", bold=True, size=11)
        rev_cell.number_format = CURRENCY_FORMAT

        ws.auto_filter.ref = f"A3:D{len(region_rows) + 3}"
        ws.freeze_panes = "A4"
        auto_width(ws)


def build_charts(wb: Workbook, data: list[dict]):
    """Add bar and line charts to a dedicated Charts sheet."""
    ws = wb.create_sheet("Charts")
    ws.sheet_properties.tabColor = COLOR_ACCENT
    ws.sheet_view.showGridLines = False

    # ---- Prepare chart data tables directly in the sheet ----

    # --- Bar chart: Revenue by Product ---
    product_totals = {}
    for r in data:
        product_totals[r["Product"]] = product_totals.get(r["Product"], 0) + r["Revenue"]

    ws.cell(row=1, column=1, value="Product")
    ws.cell(row=1, column=2, value="Revenue")
    for i, product in enumerate(PRODUCTS):
        ws.cell(row=i + 2, column=1, value=product)
        ws.cell(row=i + 2, column=2, value=product_totals[product])

    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "Total Revenue by Product"
    bar_chart.y_axis.title = "Revenue ($)"
    bar_chart.x_axis.title = "Product"
    bar_chart.style = 10
    bar_chart.width = 22
    bar_chart.height = 14

    cat_ref = Reference(ws, min_col=1, min_row=2, max_row=len(PRODUCTS) + 1)
    val_ref = Reference(ws, min_col=2, min_row=1, max_row=len(PRODUCTS) + 1)
    bar_chart.add_data(val_ref, titles_from_data=True)
    bar_chart.set_categories(cat_ref)
    bar_chart.shape = 4

    # Colour the bars with the company accent colour
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    for pt_idx in range(len(PRODUCTS)):
        pt = DataPoint(idx=pt_idx)
        pt.graphicalProperties.solidFill = COLOR_SECONDARY
        bar_chart.series[0].data_points.append(pt)

    ws.add_chart(bar_chart, "D1")

    # --- Line chart: Monthly revenue trend ---
    monthly_totals = {}
    for r in data:
        monthly_totals[r["Month"]] = monthly_totals.get(r["Month"], 0) + r["Revenue"]

    line_start_row = len(PRODUCTS) + 4
    ws.cell(row=line_start_row, column=1, value="Month")
    ws.cell(row=line_start_row, column=2, value="Revenue")
    for i, month in enumerate(MONTHS):
        ws.cell(row=line_start_row + 1 + i, column=1, value=month)
        ws.cell(row=line_start_row + 1 + i, column=2, value=monthly_totals[month])

    line_chart = LineChart()
    line_chart.title = "Monthly Revenue Trend"
    line_chart.y_axis.title = "Revenue ($)"
    line_chart.x_axis.title = "Month"
    line_chart.style = 10
    line_chart.width = 22
    line_chart.height = 14

    cat_ref2 = Reference(ws, min_col=1, min_row=line_start_row + 1,
                         max_row=line_start_row + len(MONTHS))
    val_ref2 = Reference(ws, min_col=2, min_row=line_start_row,
                         max_row=line_start_row + len(MONTHS))
    line_chart.add_data(val_ref2, titles_from_data=True)
    line_chart.set_categories(cat_ref2)

    # Style the line
    series = line_chart.series[0]
    series.graphicalProperties.line.solidFill = COLOR_ACCENT
    series.graphicalProperties.line.width = 28000  # EMUs (~2pt)
    series.smooth = True

    ws.add_chart(line_chart, "D18")

    # Hide the raw data columns (narrow them)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 2  # spacer


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    print(f"Generating {OUTPUT_FILE} ...")

    data = generate_sales_data()
    wb = Workbook()

    build_cover_sheet(wb)
    build_summary_sheet(wb, data)
    build_detail_sheet(wb, data)
    build_region_sheets(wb, data)
    build_charts(wb, data)

    # Set the active sheet to Summary on open
    wb.active = wb.sheetnames.index("Summary")

    wb.save(OUTPUT_FILE)
    print(f"Done. Saved to {OUTPUT_FILE}")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")
    print(f"  Data rows: {len(data)}")
    print(f"  Total revenue: ${sum(r['Revenue'] for r in data):,.0f}")


if __name__ == "__main__":
    main()
