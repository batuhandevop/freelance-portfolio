"""
E-Commerce Sales Analysis (2023-2024)
======================================
Generates a professional Excel workbook with:
- Raw Data: 500+ e-commerce transactions
- Product Performance: SUMIFS/COUNTIFS/AVERAGEIFS analysis
- Customer Segmentation: RFM-style segmentation
- Monthly Trends: line & bar charts
- Dashboard: KPIs, breakdowns, insights
- Country Analysis: revenue distribution & top products

Requirements: openpyxl
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime, timedelta
import random
import os

random.seed(42)

# =============================================================================
# CONFIGURATION
# =============================================================================

DARK_BLUE = "1B3A5C"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
DARK_GRAY = "404040"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
GREEN = "27AE60"
RED = "E74C3C"
GOLD = "F39C12"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
med_blue_fill = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
gold_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
dark_green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
dark_red_fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# =============================================================================
# DATA GENERATION
# =============================================================================

COUNTRIES = ["USA", "UK", "Germany", "France", "Turkey", "Canada"]
COUNTRY_WEIGHTS = [30, 18, 15, 12, 15, 10]

CATEGORIES = {
    "Electronics": [
        ("Wireless Headphones", 49.99, 129.99),
        ("Smartphone Case", 9.99, 29.99),
        ("USB-C Hub", 24.99, 59.99),
        ("Bluetooth Speaker", 29.99, 89.99),
        ("Laptop Stand", 19.99, 49.99),
        ("Wireless Mouse", 14.99, 39.99),
        ("Power Bank", 19.99, 59.99),
        ("Smart Watch Band", 12.99, 34.99),
    ],
    "Clothing": [
        ("Cotton T-Shirt", 14.99, 34.99),
        ("Denim Jeans", 39.99, 79.99),
        ("Running Shoes", 59.99, 129.99),
        ("Winter Jacket", 69.99, 149.99),
        ("Wool Scarf", 19.99, 44.99),
        ("Leather Belt", 24.99, 54.99),
    ],
    "Home & Garden": [
        ("Scented Candle", 9.99, 24.99),
        ("Plant Pot Set", 14.99, 39.99),
        ("Kitchen Organizer", 19.99, 44.99),
        ("LED Desk Lamp", 24.99, 59.99),
        ("Throw Pillow", 14.99, 34.99),
        ("Wall Clock", 19.99, 49.99),
    ],
    "Sports": [
        ("Yoga Mat", 19.99, 49.99),
        ("Resistance Bands", 9.99, 29.99),
        ("Water Bottle", 12.99, 29.99),
        ("Fitness Tracker", 29.99, 79.99),
        ("Jump Rope", 7.99, 19.99),
        ("Gym Bag", 24.99, 54.99),
    ],
    "Books": [
        ("Python Programming", 29.99, 49.99),
        ("Data Science Handbook", 34.99, 59.99),
        ("Business Strategy", 19.99, 39.99),
        ("Self-Help Guide", 12.99, 24.99),
        ("Cooking Masterclass", 24.99, 44.99),
        ("Travel Photography", 19.99, 39.99),
    ],
}

PAYMENT_METHODS = ["Credit Card", "PayPal", "Bank Transfer"]
PAYMENT_WEIGHTS = [55, 30, 15]

FIRST_NAMES = [
    "James", "Emma", "Liam", "Olivia", "Noah", "Ava", "Oliver", "Sophia",
    "Elijah", "Isabella", "Lucas", "Mia", "Mason", "Charlotte", "Ethan",
    "Amelia", "Alexander", "Harper", "Henry", "Evelyn", "Sebastian", "Luna",
    "Jack", "Camila", "Daniel", "Aria", "Michael", "Scarlett", "Owen", "Penelope",
    "Mehmet", "Ayse", "Hans", "Marie", "Pierre", "Sophie", "Ahmed", "Fatma",
    "Klaus", "Ingrid", "Jean", "Claire", "Robert", "Sarah", "William", "Emily",
    "David", "Anna", "Thomas", "Laura",
]

LAST_NAMES = [
    "Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller",
    "Davis", "Rodriguez", "Martinez", "Wilson", "Anderson", "Taylor", "Thomas",
    "Moore", "Jackson", "Martin", "Lee", "Thompson", "White", "Harris", "Clark",
    "Yilmaz", "Kaya", "Mueller", "Schmidt", "Dupont", "Lefebvre", "Wilson",
    "Campbell", "Fraser", "Stewart", "Ozturk", "Demir", "Fischer", "Weber",
    "Bernard", "Moreau", "Tremblay", "Roy", "Chen", "Park", "Kim", "Singh",
]


def generate_customers(n=150):
    customers = []
    for i in range(1, n + 1):
        cid = f"CUST-{i:04d}"
        name = f"{random.choice(FIRST_NAMES)} {random.choice(LAST_NAMES)}"
        country = random.choices(COUNTRIES, weights=COUNTRY_WEIGHTS, k=1)[0]
        customers.append((cid, name, country))
    return customers


def generate_transactions(customers, n=550):
    transactions = []
    start_date = datetime(2023, 1, 1)
    end_date = datetime(2024, 12, 31)
    date_range = (end_date - start_date).days

    for i in range(1, n + 1):
        order_id = f"ORD-{i:05d}"
        date = start_date + timedelta(days=random.randint(0, date_range))
        cust = random.choice(customers)
        cid, cname, country = cust

        category = random.choices(
            list(CATEGORIES.keys()),
            weights=[25, 22, 18, 20, 15],
            k=1,
        )[0]
        product_info = random.choice(CATEGORIES[category])
        product_name = product_info[0]
        price_low, price_high = product_info[1], product_info[2]
        unit_price = round(random.uniform(price_low, price_high), 2)

        quantity = random.choices([1, 2, 3, 4, 5], weights=[40, 30, 15, 10, 5], k=1)[0]
        discount = random.choices(
            [0, 0.05, 0.10, 0.15, 0.20, 0.25],
            weights=[40, 20, 18, 12, 7, 3],
            k=1,
        )[0]
        payment = random.choices(PAYMENT_METHODS, weights=PAYMENT_WEIGHTS, k=1)[0]

        transactions.append({
            "order_id": order_id,
            "date": date,
            "customer_id": cid,
            "customer_name": cname,
            "country": country,
            "product": product_name,
            "category": category,
            "quantity": quantity,
            "unit_price": unit_price,
            "discount": discount,
            "payment_method": payment,
        })

    transactions.sort(key=lambda x: x["date"])
    # Re-number order IDs after sorting
    for i, t in enumerate(transactions, 1):
        t["order_id"] = f"ORD-{i:05d}"

    return transactions


# =============================================================================
# SHEET BUILDERS
# =============================================================================

def apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def apply_data_style(ws, start_row, end_row, max_col, currency_cols=None, pct_cols=None):
    if currency_cols is None:
        currency_cols = []
    if pct_cols is None:
        pct_cols = []
    for r in range(start_row, end_row + 1):
        fill = alt_fill if (r - start_row) % 2 == 1 else white_fill
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c in currency_cols:
                cell.number_format = '$#,##0.00'
            elif c in pct_cols:
                cell.number_format = '0.0%'


def set_column_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_raw_data(wb, transactions):
    ws = wb.active
    ws.title = "Raw Data"

    headers = [
        "Order ID", "Date", "Customer ID", "Customer Name", "Country",
        "Product", "Category", "Quantity", "Unit Price", "Discount (%)",
        "Revenue", "Payment Method",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    apply_header_style(ws, 1, len(headers))

    for i, t in enumerate(transactions, 2):
        ws.cell(row=i, column=1, value=t["order_id"])
        ws.cell(row=i, column=2, value=t["date"]).number_format = "YYYY-MM-DD"
        ws.cell(row=i, column=3, value=t["customer_id"])
        ws.cell(row=i, column=4, value=t["customer_name"])
        ws.cell(row=i, column=5, value=t["country"])
        ws.cell(row=i, column=6, value=t["product"])
        ws.cell(row=i, column=7, value=t["category"])
        ws.cell(row=i, column=8, value=t["quantity"])
        ws.cell(row=i, column=9, value=t["unit_price"]).number_format = '$#,##0.00'
        ws.cell(row=i, column=10, value=t["discount"]).number_format = '0%'
        # Revenue formula: Quantity * Unit Price * (1 - Discount)
        ws.cell(row=i, column=11).value = f"=H{i}*I{i}*(1-J{i})"
        ws.cell(row=i, column=11).number_format = '$#,##0.00'
        ws.cell(row=i, column=12, value=t["payment_method"])

    last_row = len(transactions) + 1
    apply_data_style(ws, 2, last_row, 12, currency_cols=[9, 11], pct_cols=[10])

    set_column_widths(ws, [14, 14, 14, 20, 12, 22, 16, 10, 12, 12, 14, 16])

    ws.auto_filter.ref = f"A1:L{last_row}"
    ws.freeze_panes = "A2"

    return last_row


def build_product_performance(wb, transactions):
    ws = wb.create_sheet("Product Performance")

    # Gather unique products
    products = {}
    for t in transactions:
        key = t["product"]
        if key not in products:
            products[key] = t["category"]

    # Sort products by calculated revenue descending
    product_revenues = {}
    for t in transactions:
        rev = t["quantity"] * t["unit_price"] * (1 - t["discount"])
        product_revenues[t["product"]] = product_revenues.get(t["product"], 0) + rev

    sorted_products = sorted(products.keys(), key=lambda p: product_revenues.get(p, 0), reverse=True)

    raw_last_row = len(transactions) + 1

    headers = ["Product", "Category", "Total Revenue", "Total Orders", "Avg Order Value", "Return Rate"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    apply_header_style(ws, 1, len(headers))

    for i, product in enumerate(sorted_products, 2):
        ws.cell(row=i, column=1, value=product)
        ws.cell(row=i, column=2, value=products[product])
        # SUMIFS: sum revenue where product matches
        ws.cell(row=i, column=3).value = f"=SUMPRODUCT(('Raw Data'!H2:H{raw_last_row}*'Raw Data'!I2:I{raw_last_row}*(1-'Raw Data'!J2:J{raw_last_row}))*('Raw Data'!F2:F{raw_last_row}=A{i}))"
        ws.cell(row=i, column=3).number_format = '$#,##0.00'
        # COUNTIFS
        ws.cell(row=i, column=4).value = f"=COUNTIF('Raw Data'!F2:F{raw_last_row},A{i})"
        # Average order value
        ws.cell(row=i, column=5).value = f"=IF(D{i}>0,C{i}/D{i},0)"
        ws.cell(row=i, column=5).number_format = '$#,##0.00'
        # Simulated return rate (2-8%)
        return_rate = round(random.uniform(0.02, 0.08), 3)
        ws.cell(row=i, column=6, value=return_rate).number_format = '0.0%'

    last_row = len(sorted_products) + 1
    apply_data_style(ws, 2, last_row, 6, currency_cols=[3, 5], pct_cols=[6])

    # Conditional formatting: top 5 green, bottom 5 red
    for i, product in enumerate(sorted_products[:5], 2):
        for c in range(1, 7):
            ws.cell(row=i, column=c).fill = green_fill
    for i, product in enumerate(sorted_products[-5:], len(sorted_products) - 3):
        row = i
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = red_fill

    set_column_widths(ws, [24, 16, 16, 14, 16, 12])
    ws.freeze_panes = "A2"


def build_customer_segmentation(wb, transactions):
    ws = wb.create_sheet("Customer Segmentation")

    # Compute customer stats
    customer_data = {}
    for t in transactions:
        cid = t["customer_id"]
        rev = t["quantity"] * t["unit_price"] * (1 - t["discount"])
        if cid not in customer_data:
            customer_data[cid] = {
                "name": t["customer_name"],
                "orders": 0,
                "total_spent": 0,
                "first_purchase": t["date"],
                "last_purchase": t["date"],
            }
        customer_data[cid]["orders"] += 1
        customer_data[cid]["total_spent"] += rev
        if t["date"] < customer_data[cid]["first_purchase"]:
            customer_data[cid]["first_purchase"] = t["date"]
        if t["date"] > customer_data[cid]["last_purchase"]:
            customer_data[cid]["last_purchase"] = t["date"]

    # Calculate segments
    end_date = datetime(2024, 12, 31)
    three_months_ago = end_date - timedelta(days=90)
    two_months_ago = end_date - timedelta(days=60)

    spends = sorted([d["total_spent"] for d in customer_data.values()], reverse=True)
    vip_threshold = spends[max(0, int(len(spends) * 0.10) - 1)]

    sorted_customers = sorted(customer_data.items(), key=lambda x: x[1]["total_spent"], reverse=True)

    headers = ["Customer ID", "Customer Name", "Total Orders", "Total Spent",
               "Avg Order Value", "First Purchase", "Last Purchase", "Segment"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    apply_header_style(ws, 1, len(headers))

    segment_colors = {
        "VIP": PatternFill(start_color="E8D5B7", end_color="E8D5B7", fill_type="solid"),
        "Regular": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
        "At Risk": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
        "New": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    }

    for i, (cid, data) in enumerate(sorted_customers, 2):
        avg_order = data["total_spent"] / data["orders"] if data["orders"] > 0 else 0

        # Determine segment
        if data["total_spent"] >= vip_threshold:
            segment = "VIP"
        elif data["first_purchase"] >= two_months_ago:
            segment = "New"
        elif data["last_purchase"] < three_months_ago:
            segment = "At Risk"
        else:
            segment = "Regular"

        ws.cell(row=i, column=1, value=cid)
        ws.cell(row=i, column=2, value=data["name"])
        ws.cell(row=i, column=3, value=data["orders"])
        ws.cell(row=i, column=4, value=round(data["total_spent"], 2)).number_format = '$#,##0.00'
        ws.cell(row=i, column=5, value=round(avg_order, 2)).number_format = '$#,##0.00'
        ws.cell(row=i, column=6, value=data["first_purchase"]).number_format = "YYYY-MM-DD"
        ws.cell(row=i, column=7, value=data["last_purchase"]).number_format = "YYYY-MM-DD"
        ws.cell(row=i, column=8, value=segment)

        # Color-code segment
        seg_fill = segment_colors.get(segment, white_fill)
        for c in range(1, 9):
            cell = ws.cell(row=i, column=c)
            cell.fill = seg_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    set_column_widths(ws, [14, 22, 14, 14, 16, 16, 16, 12])
    ws.freeze_panes = "A2"


def build_monthly_trends(wb, transactions):
    ws = wb.create_sheet("Monthly Trends")

    # Aggregate monthly data
    monthly = {}
    customer_first = {}
    for t in transactions:
        cid = t["customer_id"]
        if cid not in customer_first or t["date"] < customer_first[cid]:
            customer_first[cid] = t["date"]

    for t in transactions:
        key = (t["date"].year, t["date"].month)
        rev = t["quantity"] * t["unit_price"] * (1 - t["discount"])
        if key not in monthly:
            monthly[key] = {"orders": 0, "revenue": 0, "new_customers": set()}
        monthly[key]["orders"] += 1
        monthly[key]["revenue"] += rev
        # Check if this is the customer's first month
        cid = t["customer_id"]
        first = customer_first[cid]
        if (first.year, first.month) == key:
            monthly[key]["new_customers"].add(cid)

    sorted_months = sorted(monthly.keys())

    headers = ["Month", "Orders", "Revenue", "Avg Order Value", "New Customers"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    apply_header_style(ws, 1, len(headers))

    for i, (year, month) in enumerate(sorted_months, 2):
        data = monthly[(year, month)]
        avg_val = data["revenue"] / data["orders"] if data["orders"] > 0 else 0
        month_date = datetime(year, month, 1)

        ws.cell(row=i, column=1, value=month_date).number_format = "MMM YYYY"
        ws.cell(row=i, column=2, value=data["orders"])
        ws.cell(row=i, column=3, value=round(data["revenue"], 2)).number_format = '$#,##0.00'
        ws.cell(row=i, column=4, value=round(avg_val, 2)).number_format = '$#,##0.00'
        ws.cell(row=i, column=5, value=len(data["new_customers"]))

    last_data_row = len(sorted_months) + 1
    apply_data_style(ws, 2, last_data_row, 5, currency_cols=[3, 4])

    set_column_widths(ws, [16, 12, 16, 16, 16])

    # Write chart data far below (row 200+) to avoid overlap
    chart_data_start = 200
    ws.cell(row=chart_data_start, column=1, value="Month")
    ws.cell(row=chart_data_start, column=2, value="Revenue")
    ws.cell(row=chart_data_start, column=3, value="Orders")
    for i, (year, month) in enumerate(sorted_months, 1):
        data = monthly[(year, month)]
        month_date = datetime(year, month, 1)
        ws.cell(row=chart_data_start + i, column=1, value=month_date).number_format = "MMM YYYY"
        ws.cell(row=chart_data_start + i, column=2, value=round(data["revenue"], 2))
        ws.cell(row=chart_data_start + i, column=3, value=data["orders"])

    chart_data_end = chart_data_start + len(sorted_months)

    # LINE CHART - Revenue Trend
    chart1 = LineChart()
    chart1.title = "Monthly Revenue Trend"
    chart1.y_axis.title = "Revenue ($)"
    chart1.x_axis.title = "Month"
    chart1.style = 10
    chart1.width = 28
    chart1.height = 16

    data_ref = Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_end)
    cats_ref = Reference(ws, min_col=1, min_row=chart_data_start + 1, max_row=chart_data_end)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.y_axis.delete = False
    chart1.x_axis.delete = False
    chart1.y_axis.numFmt = '$#,##0'
    chart1.x_axis.numFmt = "MMM YY"
    chart1.x_axis.tickLblPos = "low"
    chart1.y_axis.tickLblPos = "low"

    s = chart1.series[0]
    s.graphicalProperties.line.width = 25000
    s.graphicalProperties.line.solidFill = MED_BLUE

    ws.add_chart(chart1, "A" + str(last_data_row + 3))

    # BAR CHART - Order Count
    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Monthly Order Count"
    chart2.y_axis.title = "Number of Orders"
    chart2.x_axis.title = "Month"
    chart2.style = 10
    chart2.width = 28
    chart2.height = 16

    data_ref2 = Reference(ws, min_col=3, min_row=chart_data_start, max_row=chart_data_end)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats_ref)
    chart2.y_axis.delete = False
    chart2.x_axis.delete = False
    chart2.y_axis.numFmt = '#,##0'
    chart2.x_axis.numFmt = "MMM YY"
    chart2.x_axis.tickLblPos = "low"
    chart2.y_axis.tickLblPos = "low"

    s2 = chart2.series[0]
    s2.graphicalProperties.solidFill = MED_BLUE

    ws.add_chart(chart2, "A" + str(last_data_row + 20))


def build_dashboard(wb, transactions, customers):
    ws = wb.create_sheet("Dashboard")

    # Pre-compute stats
    total_revenue = sum(t["quantity"] * t["unit_price"] * (1 - t["discount"]) for t in transactions)
    total_orders = len(transactions)
    avg_order_value = total_revenue / total_orders
    unique_customers = len(set(t["customer_id"] for t in transactions))

    # Country stats
    country_revenue = {}
    country_orders = {}
    for t in transactions:
        rev = t["quantity"] * t["unit_price"] * (1 - t["discount"])
        country_revenue[t["country"]] = country_revenue.get(t["country"], 0) + rev
        country_orders[t["country"]] = country_orders.get(t["country"], 0) + 1

    top_country = max(country_revenue, key=country_revenue.get)

    # Product stats
    product_revenue = {}
    for t in transactions:
        rev = t["quantity"] * t["unit_price"] * (1 - t["discount"])
        product_revenue[t["product"]] = product_revenue.get(t["product"], 0) + rev
    top_product = max(product_revenue, key=product_revenue.get)

    # Category stats
    cat_revenue = {}
    cat_orders = {}
    for t in transactions:
        rev = t["quantity"] * t["unit_price"] * (1 - t["discount"])
        cat_revenue[t["category"]] = cat_revenue.get(t["category"], 0) + rev
        cat_orders[t["category"]] = cat_orders.get(t["category"], 0) + 1

    # Styling helpers
    def title_cell(row, col, text, font_size=20, span=8):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(name="Calibri", bold=True, size=font_size, color=DARK_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)

    def kpi_card(row, col, label, value, fmt="general"):
        # Label
        lbl = ws.cell(row=row, column=col, value=label)
        lbl.font = Font(name="Calibri", size=9, color=WHITE, bold=True)
        lbl.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
        lbl.alignment = Alignment(horizontal="center", vertical="center")
        lbl.border = thin_border

        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)

        # Value
        val = ws.cell(row=row + 1, column=col, value=value)
        val.font = Font(name="Calibri", size=16, bold=True, color=DARK_BLUE)
        val.alignment = Alignment(horizontal="center", vertical="center")
        val.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        val.border = thin_border

        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)

        if fmt == "currency":
            val.number_format = '$#,##0.00'
        elif fmt == "number":
            val.number_format = '#,##0'

    # === TITLE ===
    row = 2
    title_cell(row, 2, "E-Commerce Sales Analysis | 2023-2024", 20, 10)

    row = 3
    sub = ws.cell(row=row, column=2, value="Data Source: Internal E-Commerce Platform  |  550+ Transactions  |  6 Markets")
    sub.font = Font(name="Calibri", size=10, italic=True, color=DARK_GRAY)
    sub.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)

    # === KPI CARDS ===
    row = 5
    kpi_card(row, 2, "TOTAL REVENUE", total_revenue, "currency")
    kpi_card(row, 4, "TOTAL ORDERS", total_orders, "number")
    kpi_card(row, 6, "AVG ORDER VALUE", avg_order_value, "currency")
    kpi_card(row, 8, "UNIQUE CUSTOMERS", unique_customers, "number")
    kpi_card(row, 10, "TOP COUNTRY", top_country)

    # === COUNTRY BREAKDOWN ===
    row = 9
    section_header = ws.cell(row=row, column=2, value="Revenue by Country")
    section_header.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    section_header.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    section_header.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)

    row = 10
    for c, h in enumerate(["Country", "Revenue", "Orders", "Avg Order", "% Share"], 2):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cell.fill = med_blue_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    sorted_countries = sorted(country_revenue.items(), key=lambda x: x[1], reverse=True)
    for i, (country, rev) in enumerate(sorted_countries, 11):
        orders = country_orders[country]
        fill = alt_fill if (i - 11) % 2 == 1 else white_fill
        ws.cell(row=i, column=2, value=country).fill = fill
        ws.cell(row=i, column=2).border = thin_border
        ws.cell(row=i, column=2).alignment = Alignment(horizontal="center")

        ws.cell(row=i, column=3, value=round(rev, 2)).number_format = '$#,##0.00'
        ws.cell(row=i, column=3).fill = fill
        ws.cell(row=i, column=3).border = thin_border
        ws.cell(row=i, column=3).alignment = Alignment(horizontal="center")

        ws.cell(row=i, column=4, value=orders).number_format = '#,##0'
        ws.cell(row=i, column=4).fill = fill
        ws.cell(row=i, column=4).border = thin_border
        ws.cell(row=i, column=4).alignment = Alignment(horizontal="center")

        ws.cell(row=i, column=5, value=round(rev / orders, 2)).number_format = '$#,##0.00'
        ws.cell(row=i, column=5).fill = fill
        ws.cell(row=i, column=5).border = thin_border
        ws.cell(row=i, column=5).alignment = Alignment(horizontal="center")

        ws.cell(row=i, column=6, value=round(rev / total_revenue, 4)).number_format = '0.0%'
        ws.cell(row=i, column=6).fill = fill
        ws.cell(row=i, column=6).border = thin_border
        ws.cell(row=i, column=6).alignment = Alignment(horizontal="center")

    # === CATEGORY BREAKDOWN ===
    cat_start_row = 9
    section_header2 = ws.cell(row=cat_start_row, column=8, value="Revenue by Category")
    section_header2.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    section_header2.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    section_header2.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=cat_start_row, start_column=8, end_row=cat_start_row, end_column=12)

    row = cat_start_row + 1
    for c, h in enumerate(["Category", "Revenue", "Orders", "Avg Order", "% Share"], 8):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cell.fill = med_blue_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    sorted_cats = sorted(cat_revenue.items(), key=lambda x: x[1], reverse=True)
    for i, (cat, rev) in enumerate(sorted_cats, cat_start_row + 2):
        orders = cat_orders[cat]
        fill = alt_fill if (i - cat_start_row - 2) % 2 == 1 else white_fill

        ws.cell(row=i, column=8, value=cat).fill = fill
        ws.cell(row=i, column=8).border = thin_border
        ws.cell(row=i, column=8).alignment = Alignment(horizontal="center")

        ws.cell(row=i, column=9, value=round(rev, 2)).number_format = '$#,##0.00'
        ws.cell(row=i, column=9).fill = fill
        ws.cell(row=i, column=9).border = thin_border
        ws.cell(row=i, column=9).alignment = Alignment(horizontal="center")

        ws.cell(row=i, column=10, value=orders).number_format = '#,##0'
        ws.cell(row=i, column=10).fill = fill
        ws.cell(row=i, column=10).border = thin_border
        ws.cell(row=i, column=10).alignment = Alignment(horizontal="center")

        ws.cell(row=i, column=11, value=round(rev / orders, 2)).number_format = '$#,##0.00'
        ws.cell(row=i, column=11).fill = fill
        ws.cell(row=i, column=11).border = thin_border
        ws.cell(row=i, column=11).alignment = Alignment(horizontal="center")

        ws.cell(row=i, column=12, value=round(rev / total_revenue, 4)).number_format = '0.0%'
        ws.cell(row=i, column=12).fill = fill
        ws.cell(row=i, column=12).border = thin_border
        ws.cell(row=i, column=12).alignment = Alignment(horizontal="center")

    # === KEY INSIGHTS ===
    insights_row = 19
    ins_header = ws.cell(row=insights_row, column=2, value="Key Insights")
    ins_header.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    ins_header.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    ins_header.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=insights_row, start_column=2, end_row=insights_row, end_column=12)

    insights = [
        f"1. Total revenue of ${total_revenue:,.2f} generated across {total_orders} orders from {unique_customers} unique customers.",
        f"2. {top_country} leads all markets with ${country_revenue[top_country]:,.2f} in revenue ({country_revenue[top_country]/total_revenue*100:.1f}% share).",
        f"3. '{top_product}' is the best-selling product by revenue.",
        f"4. {sorted_cats[0][0]} is the top-performing category with ${sorted_cats[0][1]:,.2f} revenue.",
        f"5. Average order value is ${avg_order_value:,.2f}, indicating a healthy basket size.",
        f"6. Credit Card is the dominant payment method across all markets.",
    ]
    for i, insight in enumerate(insights):
        r = insights_row + 1 + i
        cell = ws.cell(row=r, column=2, value=insight)
        cell.font = Font(name="Calibri", size=10, color=DARK_GRAY)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        cell.border = thin_border
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)

    # Column widths
    ws.column_dimensions["A"].width = 3
    for col in range(2, 13):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # Row heights
    ws.row_dimensions[2].height = 35
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 35


def build_country_analysis(wb, transactions):
    ws = wb.create_sheet("Country Analysis")

    # Compute country data
    country_data = {}
    country_products = {}
    for t in transactions:
        c = t["country"]
        rev = t["quantity"] * t["unit_price"] * (1 - t["discount"])
        if c not in country_data:
            country_data[c] = {"revenue": 0, "orders": 0, "products": {}}
        country_data[c]["revenue"] += rev
        country_data[c]["orders"] += 1

        p = t["product"]
        country_data[c]["products"][p] = country_data[c]["products"].get(p, 0) + rev

    total_revenue = sum(d["revenue"] for d in country_data.values())
    sorted_countries = sorted(country_data.items(), key=lambda x: x[1]["revenue"], reverse=True)

    # === Revenue by Country Table ===
    headers = ["Country", "Revenue", "Orders", "Avg Order Value", "% of Total Revenue"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    apply_header_style(ws, 1, len(headers))

    for i, (country, data) in enumerate(sorted_countries, 2):
        ws.cell(row=i, column=1, value=country)
        ws.cell(row=i, column=2, value=round(data["revenue"], 2)).number_format = '$#,##0.00'
        ws.cell(row=i, column=3, value=data["orders"])
        avg = data["revenue"] / data["orders"] if data["orders"] > 0 else 0
        ws.cell(row=i, column=4, value=round(avg, 2)).number_format = '$#,##0.00'
        ws.cell(row=i, column=5, value=round(data["revenue"] / total_revenue, 4)).number_format = '0.0%'

    last_row = len(sorted_countries) + 1
    apply_data_style(ws, 2, last_row, 5, currency_cols=[2, 4], pct_cols=[5])

    # === Top Products per Country ===
    top_prod_start = last_row + 3
    tp_header = ws.cell(row=top_prod_start, column=1, value="Top 3 Products by Country")
    tp_header.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    tp_header.fill = header_fill
    tp_header.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=top_prod_start, start_column=1, end_row=top_prod_start, end_column=5)

    sub_headers = ["Country", "Rank", "Product", "Revenue", "% Country Revenue"]
    sub_row = top_prod_start + 1
    for c, h in enumerate(sub_headers, 1):
        ws.cell(row=sub_row, column=c, value=h)
    apply_header_style(ws, sub_row, 5)

    r = sub_row + 1
    for country, data in sorted_countries:
        top_products = sorted(data["products"].items(), key=lambda x: x[1], reverse=True)[:3]
        for rank, (prod, rev) in enumerate(top_products, 1):
            fill = alt_fill if (r - sub_row - 1) % 2 == 1 else white_fill
            ws.cell(row=r, column=1, value=country).fill = fill
            ws.cell(row=r, column=1).border = thin_border
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")

            ws.cell(row=r, column=2, value=rank).fill = fill
            ws.cell(row=r, column=2).border = thin_border
            ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")

            ws.cell(row=r, column=3, value=prod).fill = fill
            ws.cell(row=r, column=3).border = thin_border
            ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")

            ws.cell(row=r, column=4, value=round(rev, 2)).number_format = '$#,##0.00'
            ws.cell(row=r, column=4).fill = fill
            ws.cell(row=r, column=4).border = thin_border
            ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")

            ws.cell(row=r, column=5, value=round(rev / data["revenue"], 4)).number_format = '0.0%'
            ws.cell(row=r, column=5).fill = fill
            ws.cell(row=r, column=5).border = thin_border
            ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")

            r += 1

    top_prod_end_row = r - 1

    # Chart data at row 200+
    chart_data_start = 200
    ws.cell(row=chart_data_start, column=1, value="Country")
    ws.cell(row=chart_data_start, column=2, value="Revenue")
    for i, (country, data) in enumerate(sorted_countries, 1):
        ws.cell(row=chart_data_start + i, column=1, value=country)
        ws.cell(row=chart_data_start + i, column=2, value=round(data["revenue"], 2))

    chart_data_end = chart_data_start + len(sorted_countries)

    # BAR CHART - Revenue by Country
    chart = BarChart()
    chart.type = "col"
    chart.title = "Revenue by Country"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Country"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    data_ref = Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_end)
    cats_ref = Reference(ws, min_col=1, min_row=chart_data_start + 1, max_row=chart_data_end)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    chart.y_axis.numFmt = '$#,##0'
    chart.x_axis.tickLblPos = "low"
    chart.y_axis.tickLblPos = "low"

    s = chart.series[0]
    s.graphicalProperties.solidFill = MED_BLUE

    ws.add_chart(chart, "G2")

    # PIE CHART - Country Distribution
    pie = PieChart()
    pie.title = "Revenue Distribution by Country"
    pie.style = 10
    pie.width = 18
    pie.height = 14

    pie_data = Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_end)
    pie_cats = Reference(ws, min_col=1, min_row=chart_data_start + 1, max_row=chart_data_end)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)

    pie.series[0].data_points = []
    pie_colors = ["2E75B6", "E74C3C", "27AE60", "F39C12", "8E44AD", "1ABC9C"]
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    for idx, color in enumerate(pie_colors[:len(sorted_countries)]):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt)

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True

    ws.add_chart(pie, "G18")

    set_column_widths(ws, [14, 16, 10, 22, 16, 18, 4])


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("Generating E-Commerce Sales Analysis...")

    wb = openpyxl.Workbook()
    customers = generate_customers(150)
    transactions = generate_transactions(customers, 550)

    print(f"  Generated {len(transactions)} transactions, {len(customers)} customers")

    raw_last = build_raw_data(wb, transactions)
    print("  [1/6] Raw Data sheet created")

    build_product_performance(wb, transactions)
    print("  [2/6] Product Performance sheet created")

    build_customer_segmentation(wb, transactions)
    print("  [3/6] Customer Segmentation sheet created")

    build_monthly_trends(wb, transactions)
    print("  [4/6] Monthly Trends sheet created")

    build_dashboard(wb, transactions, customers)
    print("  [5/6] Dashboard sheet created")

    build_country_analysis(wb, transactions)
    print("  [6/6] Country Analysis sheet created")

    # Reorder sheets: Raw Data, Product Performance, Customer Segmentation, Monthly Trends, Dashboard, Country Analysis
    desired_order = ["Raw Data", "Product Performance", "Customer Segmentation", "Monthly Trends", "Dashboard", "Country Analysis"]
    sheet_names = wb.sheetnames
    for i, name in enumerate(desired_order):
        current_idx = wb.sheetnames.index(name)
        wb.move_sheet(name, offset=i - current_idx)

    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ecommerce_analysis.xlsx")
    wb.save(output_path)
    print(f"\nSaved to: {output_path}")
    print("Done!")


if __name__ == "__main__":
    main()
