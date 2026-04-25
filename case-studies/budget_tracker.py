"""
Financial Budget Tracker — Case Study
Generates budget_tracker.xlsx with 6 professional sheets.
"""

import random
import datetime
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import (
    LineChart, BarChart, PieChart, Reference
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

random.seed(42)

# ── colour palette ──────────────────────────────────────────────
DARK_BLUE   = "1B3A5C"
MED_BLUE    = "2E75B6"
GREEN       = "548235"
RED         = "C00000"
LIGHT_GRAY  = "F2F2F2"
WHITE       = "FFFFFF"
DARK_GRAY   = "404040"
LIGHT_BLUE  = "D6E4F0"
LIGHT_GREEN = "E2EFDA"
LIGHT_RED   = "FCE4EC"
BLUE_KPI    = "1F4E79"

header_font  = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill  = PatternFill("solid", fgColor=DARK_BLUE)
alt_fill     = PatternFill("solid", fgColor=LIGHT_GRAY)
green_fill   = PatternFill("solid", fgColor=LIGHT_GREEN)
red_fill     = PatternFill("solid", fgColor=LIGHT_RED)
thin_border  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
currency_fmt = '$#,##0.00'
pct_fmt      = '0.0%'

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

# ── category definitions ────────────────────────────────────────
EXPENSE_CATS = {
    "Housing":        ["Rent", "Utilities", "Internet", "Insurance"],
    "Food":           ["Groceries", "Dining Out", "Coffee"],
    "Transportation": ["Fuel", "Public Transit", "Car Maintenance"],
    "Health":         ["Gym", "Medical", "Pharmacy"],
    "Entertainment":  ["Streaming", "Games", "Events"],
    "Shopping":       ["Clothing", "Electronics", "Home"],
    "Education":      ["Courses", "Books"],
    "Savings":        ["Emergency Fund", "Retirement"],
}

INCOME_CATS = {
    "Salary":      ["Monthly Salary"],
    "Freelance":   ["Web Development", "Consulting", "Design Work"],
    "Investments": ["Dividends", "Interest"],
    "Other":       ["Bonus", "Gift", "Refund"],
}

PAYMENT_METHODS = ["Cash", "Credit Card", "Bank Transfer", "Direct Debit"]

# monthly budgets per expense category
BUDGETS = {
    "Housing":        1800,
    "Food":           600,
    "Transportation": 350,
    "Health":         200,
    "Entertainment":  150,
    "Shopping":       300,
    "Education":      150,
    "Savings":        500,
}

# ── helper: style a header row ──────────────────────────────────
def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def style_data_area(ws, start_row, end_row, max_col, currency_cols=None, pct_cols=None):
    currency_cols = currency_cols or []
    pct_cols = pct_cols or []
    for r in range(start_row, end_row + 1):
        fill = alt_fill if (r - start_row) % 2 == 1 else PatternFill()
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill.fgColor and fill.fgColor.rgb != "00000000":
                cell.fill = fill
            if c in currency_cols:
                cell.number_format = currency_fmt
            if c in pct_cols:
                cell.number_format = pct_fmt


def auto_width(ws, max_col, min_width=12, max_width=22):
    for c in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = min(max(min_width, 14), max_width)


# ── generate transactions ──────────────────────────────────────
def generate_transactions():
    txns = []

    for month in range(1, 13):
        # -- salary 1st
        txns.append({
            "date": datetime.date(2024, month, 1),
            "type": "Income", "category": "Salary",
            "subcategory": "Monthly Salary",
            "description": "Monthly salary deposit",
            "amount": round(random.uniform(5200, 5600), 2),
            "payment": "Bank Transfer",
        })

        # -- freelance (1-3 per month)
        for _ in range(random.randint(1, 3)):
            sub = random.choice(INCOME_CATS["Freelance"])
            txns.append({
                "date": datetime.date(2024, month, random.randint(1, 28)),
                "type": "Income", "category": "Freelance",
                "subcategory": sub,
                "description": f"Freelance {sub.lower()} project",
                "amount": round(random.uniform(300, 2500), 2),
                "payment": "Bank Transfer",
            })

        # -- investments (quarterly)
        if month in (3, 6, 9, 12):
            txns.append({
                "date": datetime.date(2024, month, 15),
                "type": "Income", "category": "Investments",
                "subcategory": "Dividends",
                "description": "Quarterly dividend payment",
                "amount": round(random.uniform(150, 400), 2),
                "payment": "Bank Transfer",
            })
            txns.append({
                "date": datetime.date(2024, month, 15),
                "type": "Income", "category": "Investments",
                "subcategory": "Interest",
                "description": "Savings account interest",
                "amount": round(random.uniform(20, 80), 2),
                "payment": "Bank Transfer",
            })

        # -- other income (random)
        if random.random() < 0.3:
            sub = random.choice(INCOME_CATS["Other"])
            txns.append({
                "date": datetime.date(2024, month, random.randint(1, 28)),
                "type": "Income", "category": "Other",
                "subcategory": sub,
                "description": f"{sub} received",
                "amount": round(random.uniform(50, 500), 2),
                "payment": random.choice(["Cash", "Bank Transfer"]),
            })

        # -- rent 5th
        txns.append({
            "date": datetime.date(2024, month, 5),
            "type": "Expense", "category": "Housing",
            "subcategory": "Rent",
            "description": "Monthly rent payment",
            "amount": round(random.uniform(1200, 1300), 2),
            "payment": "Bank Transfer",
        })

        # -- utilities
        txns.append({
            "date": datetime.date(2024, month, random.randint(10, 15)),
            "type": "Expense", "category": "Housing",
            "subcategory": "Utilities",
            "description": "Electric & water bill",
            "amount": round(random.uniform(100, 220) * (1.3 if month in (1,2,7,8) else 1), 2),
            "payment": "Direct Debit",
        })

        # -- internet
        txns.append({
            "date": datetime.date(2024, month, 1),
            "type": "Expense", "category": "Housing",
            "subcategory": "Internet",
            "description": "Internet service",
            "amount": 79.99,
            "payment": "Direct Debit",
        })

        # -- insurance (quarterly)
        if month in (1, 4, 7, 10):
            txns.append({
                "date": datetime.date(2024, month, 20),
                "type": "Expense", "category": "Housing",
                "subcategory": "Insurance",
                "description": "Renters insurance quarterly",
                "amount": round(random.uniform(180, 220), 2),
                "payment": "Direct Debit",
            })

        # -- groceries 4-6 times per month
        for _ in range(random.randint(4, 6)):
            txns.append({
                "date": datetime.date(2024, month, random.randint(1, 28)),
                "type": "Expense", "category": "Food",
                "subcategory": "Groceries",
                "description": random.choice([
                    "Weekly groceries", "Grocery run", "Supermarket",
                    "Farmers market", "Bulk grocery shopping"
                ]),
                "amount": round(random.uniform(40, 130), 2),
                "payment": random.choice(["Credit Card", "Cash"]),
            })

        # -- dining out 2-4
        for _ in range(random.randint(2, 4)):
            txns.append({
                "date": datetime.date(2024, month, random.randint(1, 28)),
                "type": "Expense", "category": "Food",
                "subcategory": "Dining Out",
                "description": random.choice([
                    "Restaurant dinner", "Lunch out", "Pizza night",
                    "Brunch with friends", "Takeout order"
                ]),
                "amount": round(random.uniform(15, 75), 2),
                "payment": "Credit Card",
            })

        # -- coffee
        for _ in range(random.randint(3, 8)):
            txns.append({
                "date": datetime.date(2024, month, random.randint(1, 28)),
                "type": "Expense", "category": "Food",
                "subcategory": "Coffee",
                "description": "Coffee shop",
                "amount": round(random.uniform(4, 8), 2),
                "payment": random.choice(["Cash", "Credit Card"]),
            })

        # -- fuel
        for _ in range(random.randint(2, 4)):
            txns.append({
                "date": datetime.date(2024, month, random.randint(1, 28)),
                "type": "Expense", "category": "Transportation",
                "subcategory": "Fuel",
                "description": "Gas station fill-up",
                "amount": round(random.uniform(35, 70), 2),
                "payment": "Credit Card",
            })

        # -- public transit
        if random.random() < 0.6:
            txns.append({
                "date": datetime.date(2024, month, 1),
                "type": "Expense", "category": "Transportation",
                "subcategory": "Public Transit",
                "description": "Monthly transit pass",
                "amount": round(random.uniform(50, 90), 2),
                "payment": "Credit Card",
            })

        # -- car maintenance (every few months)
        if random.random() < 0.25:
            txns.append({
                "date": datetime.date(2024, month, random.randint(1, 28)),
                "type": "Expense", "category": "Transportation",
                "subcategory": "Car Maintenance",
                "description": random.choice(["Oil change", "Tire rotation", "Car wash", "Brake pads"]),
                "amount": round(random.uniform(30, 300), 2),
                "payment": "Credit Card",
            })

        # -- gym
        txns.append({
            "date": datetime.date(2024, month, 1),
            "type": "Expense", "category": "Health",
            "subcategory": "Gym",
            "description": "Gym membership",
            "amount": 49.99,
            "payment": "Direct Debit",
        })

        # -- medical/pharmacy
        if random.random() < 0.3:
            txns.append({
                "date": datetime.date(2024, month, random.randint(1, 28)),
                "type": "Expense", "category": "Health",
                "subcategory": random.choice(["Medical", "Pharmacy"]),
                "description": random.choice(["Doctor visit", "Prescription", "Pharmacy purchase"]),
                "amount": round(random.uniform(20, 200), 2),
                "payment": random.choice(["Credit Card", "Cash"]),
            })

        # -- streaming
        txns.append({
            "date": datetime.date(2024, month, 1),
            "type": "Expense", "category": "Entertainment",
            "subcategory": "Streaming",
            "description": "Streaming subscriptions",
            "amount": round(15.99 + 9.99 + 7.99, 2),  # Netflix+Spotify+other
            "payment": "Credit Card",
        })

        # -- games/events (occasional)
        if random.random() < 0.4:
            txns.append({
                "date": datetime.date(2024, month, random.randint(1, 28)),
                "type": "Expense", "category": "Entertainment",
                "subcategory": random.choice(["Games", "Events"]),
                "description": random.choice(["Video game purchase", "Concert tickets", "Movie night", "Sports event"]),
                "amount": round(random.uniform(15, 120), 2),
                "payment": "Credit Card",
            })

        # -- shopping (seasonal variation)
        shop_count = random.randint(1, 3)
        if month in (11, 12):  # holiday season
            shop_count += 2
        for _ in range(shop_count):
            sub = random.choice(EXPENSE_CATS["Shopping"])
            txns.append({
                "date": datetime.date(2024, month, random.randint(1, 28)),
                "type": "Expense", "category": "Shopping",
                "subcategory": sub,
                "description": random.choice([
                    f"{sub} purchase", f"Online {sub.lower()} order",
                    f"{sub} store visit"
                ]),
                "amount": round(random.uniform(20, 250) * (1.4 if month in (11,12) else 1), 2),
                "payment": random.choice(["Credit Card", "Cash"]),
            })

        # -- education
        if random.random() < 0.35:
            sub = random.choice(EXPENSE_CATS["Education"])
            txns.append({
                "date": datetime.date(2024, month, random.randint(1, 28)),
                "type": "Expense", "category": "Education",
                "subcategory": sub,
                "description": random.choice([
                    "Online course", "Udemy course", "Technical book",
                    "Workshop registration"
                ]),
                "amount": round(random.uniform(10, 200), 2),
                "payment": "Credit Card",
            })

        # -- savings
        txns.append({
            "date": datetime.date(2024, month, 1),
            "type": "Expense", "category": "Savings",
            "subcategory": "Emergency Fund",
            "description": "Emergency fund contribution",
            "amount": round(random.uniform(200, 350), 2),
            "payment": "Bank Transfer",
        })
        txns.append({
            "date": datetime.date(2024, month, 1),
            "type": "Expense", "category": "Savings",
            "subcategory": "Retirement",
            "description": "Retirement account contribution",
            "amount": round(random.uniform(200, 350), 2),
            "payment": "Bank Transfer",
        })

    txns.sort(key=lambda t: t["date"])
    return txns


# ── SHEET 1: Transactions ──────────────────────────────────────
def build_transactions_sheet(wb, txns):
    ws = wb.active
    ws.title = "Transactions"
    headers = ["Date", "Type", "Category", "Subcategory",
               "Description", "Amount", "Payment Method"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    for i, t in enumerate(txns, 2):
        ws.cell(row=i, column=1, value=t["date"]).number_format = "YYYY-MM-DD"
        ws.cell(row=i, column=2, value=t["type"])
        ws.cell(row=i, column=3, value=t["category"])
        ws.cell(row=i, column=4, value=t["subcategory"])
        ws.cell(row=i, column=5, value=t["description"])
        ws.cell(row=i, column=6, value=t["amount"]).number_format = currency_fmt
        ws.cell(row=i, column=7, value=t["payment"])

    last = len(txns) + 1
    style_data_area(ws, 2, last, len(headers), currency_cols=[6])

    ws.auto_filter.ref = f"A1:G{last}"
    ws.freeze_panes = "A2"

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 18
    return ws


# ── SHEET 2: Monthly Budget ────────────────────────────────────
def build_monthly_budget(wb, txns):
    ws = wb.create_sheet("Monthly Budget")
    cats = list(BUDGETS.keys())

    # compute actuals
    actuals = {}
    for t in txns:
        if t["type"] == "Expense":
            key = (t["category"], t["date"].month)
            actuals[key] = actuals.get(key, 0) + t["amount"]

    # header row
    headers = ["Category"]
    for m in MONTHS:
        headers += [f"{m[:3]} Budget", f"{m[:3]} Actual", f"{m[:3]} Variance", f"{m[:3]} Var %"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    for i, cat in enumerate(cats, 2):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=1).font = Font(bold=True)
        for m_idx in range(12):
            base = 2 + m_idx * 4
            budget = BUDGETS[cat]
            actual = round(actuals.get((cat, m_idx + 1), 0), 2)
            variance = budget - actual
            var_pct = variance / budget if budget else 0

            ws.cell(row=i, column=base, value=budget).number_format = currency_fmt
            ws.cell(row=i, column=base + 1, value=actual).number_format = currency_fmt
            ws.cell(row=i, column=base + 2, value=variance).number_format = currency_fmt
            pct_cell = ws.cell(row=i, column=base + 3, value=var_pct)
            pct_cell.number_format = pct_fmt

            # conditional coloring
            if variance >= 0:
                ws.cell(row=i, column=base + 2).fill = PatternFill("solid", fgColor=LIGHT_GREEN)
                pct_cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
            else:
                ws.cell(row=i, column=base + 2).fill = PatternFill("solid", fgColor=LIGHT_RED)
                pct_cell.fill = PatternFill("solid", fgColor=LIGHT_RED)

    # totals row
    total_row = len(cats) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color=WHITE)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=DARK_BLUE)
    for m_idx in range(12):
        base = 2 + m_idx * 4
        budget_total = sum(BUDGETS.values())
        actual_total = sum(actuals.get((cat, m_idx + 1), 0) for cat in cats)
        variance_total = budget_total - actual_total
        var_pct_total = variance_total / budget_total if budget_total else 0

        for dc, val, fmt in [(0, budget_total, currency_fmt), (1, actual_total, currency_fmt),
                             (2, variance_total, currency_fmt), (3, var_pct_total, pct_fmt)]:
            cell = ws.cell(row=total_row, column=base + dc, value=round(val, 2))
            cell.number_format = fmt
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    last_col = len(headers)
    style_data_area(ws, 2, total_row, last_col,
                    currency_cols=[c for c in range(2, last_col + 1) if (c - 2) % 4 in (0, 1, 2)],
                    pct_cols=[c for c in range(2, last_col + 1) if (c - 2) % 4 == 3])

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 18
    for c in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    return ws


# ── SHEET 3: Monthly Summary ───────────────────────────────────
def build_monthly_summary(wb, txns):
    ws = wb.create_sheet("Monthly Summary")

    # compute
    monthly_income = [0] * 12
    monthly_expense = [0] * 12
    for t in txns:
        m = t["date"].month - 1
        if t["type"] == "Income":
            monthly_income[m] += t["amount"]
        else:
            monthly_expense[m] += t["amount"]

    headers = ["Month", "Total Income", "Total Expenses", "Net Savings", "Savings Rate (%)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    cumulative = 0
    for i in range(12):
        r = i + 2
        inc = round(monthly_income[i], 2)
        exp = round(monthly_expense[i], 2)
        net = round(inc - exp, 2)
        rate = net / inc if inc else 0
        cumulative += net

        ws.cell(row=r, column=1, value=MONTHS[i])
        ws.cell(row=r, column=2, value=inc).number_format = currency_fmt
        ws.cell(row=r, column=3, value=exp).number_format = currency_fmt
        ws.cell(row=r, column=4, value=net).number_format = currency_fmt
        ws.cell(row=r, column=5, value=rate).number_format = pct_fmt

    # cumulative row
    r = 14
    ws.cell(row=r, column=1, value="Cumulative Savings").font = Font(bold=True)
    ws.cell(row=r, column=4, value=round(cumulative, 2)).number_format = currency_fmt
    ws.cell(row=r, column=4).font = Font(bold=True, color=GREEN)

    style_data_area(ws, 2, 13, 5, currency_cols=[2, 3, 4], pct_cols=[5])
    auto_width(ws, 5)

    # ── chart data at row 200+ ──────────────────────────────────
    data_start = 200
    ws.cell(row=data_start, column=1, value="Month")
    ws.cell(row=data_start, column=2, value="Income")
    ws.cell(row=data_start, column=3, value="Expenses")
    ws.cell(row=data_start, column=4, value="Net Savings")
    for i in range(12):
        r = data_start + 1 + i
        ws.cell(row=r, column=1, value=MONTHS[i])
        ws.cell(row=r, column=2, value=round(monthly_income[i], 2))
        ws.cell(row=r, column=3, value=round(monthly_expense[i], 2))
        ws.cell(row=r, column=4, value=round(monthly_income[i] - monthly_expense[i], 2))

    # line chart: income vs expenses
    line = LineChart()
    line.title = "Income vs Expenses (2024)"
    line.y_axis.title = "Amount ($)"
    line.x_axis.title = "Month"
    line.y_axis.delete = False
    line.x_axis.delete = False
    line.y_axis.numFmt = '$#,##0'
    line.width = 20
    line.height = 12

    cats_ref = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_start + 12)
    inc_ref  = Reference(ws, min_col=2, min_row=data_start, max_row=data_start + 12)
    exp_ref  = Reference(ws, min_col=3, min_row=data_start, max_row=data_start + 12)

    line.add_data(inc_ref, titles_from_data=True)
    line.add_data(exp_ref, titles_from_data=True)
    line.set_categories(cats_ref)
    line.series[0].graphicalProperties.line.solidFill = GREEN
    line.series[1].graphicalProperties.line.solidFill = RED
    ws.add_chart(line, "A17")

    # bar chart: net savings
    bar = BarChart()
    bar.type = "col"
    bar.title = "Monthly Net Savings (2024)"
    bar.y_axis.title = "Amount ($)"
    bar.x_axis.title = "Month"
    bar.y_axis.delete = False
    bar.x_axis.delete = False
    bar.y_axis.numFmt = '$#,##0'
    bar.width = 20
    bar.height = 12

    net_ref = Reference(ws, min_col=4, min_row=data_start, max_row=data_start + 12)
    bar.add_data(net_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.series[0].graphicalProperties.solidFill = GREEN
    ws.add_chart(bar, "A34")

    return ws


# ── SHEET 4: Category Breakdown ────────────────────────────────
def build_category_breakdown(wb, txns):
    ws = wb.create_sheet("Category Breakdown")

    # compute
    cat_totals = {}
    for t in txns:
        if t["type"] == "Expense":
            cat_totals[t["category"]] = cat_totals.get(t["category"], 0) + t["amount"]
    grand_total = sum(cat_totals.values())
    sorted_cats = sorted(cat_totals.keys(), key=lambda c: cat_totals[c], reverse=True)

    headers = ["Category", "Total Spent", "% of Total", "Monthly Average",
               "Annual Budget", "Over/Under"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    for i, cat in enumerate(sorted_cats, 2):
        total = round(cat_totals[cat], 2)
        pct = total / grand_total if grand_total else 0
        avg = round(total / 12, 2)
        budget = BUDGETS.get(cat, 0) * 12
        over_under = budget - total

        ws.cell(row=i, column=1, value=cat).font = Font(bold=True)
        ws.cell(row=i, column=2, value=total).number_format = currency_fmt
        ws.cell(row=i, column=3, value=pct).number_format = pct_fmt
        ws.cell(row=i, column=4, value=avg).number_format = currency_fmt
        ws.cell(row=i, column=5, value=budget).number_format = currency_fmt
        cell = ws.cell(row=i, column=6, value=round(over_under, 2))
        cell.number_format = currency_fmt
        if over_under >= 0:
            cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
            cell.font = Font(color=GREEN)
        else:
            cell.fill = PatternFill("solid", fgColor=LIGHT_RED)
            cell.font = Font(color=RED)

    last = len(sorted_cats) + 1
    style_data_area(ws, 2, last, 6, currency_cols=[2, 4, 5, 6], pct_cols=[3])
    auto_width(ws, 6)

    # chart data at row 200+
    data_start = 200
    ws.cell(row=data_start, column=1, value="Category")
    ws.cell(row=data_start, column=2, value="Total Spent")
    ws.cell(row=data_start, column=3, value="Budget")
    for i, cat in enumerate(sorted_cats):
        r = data_start + 1 + i
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=round(cat_totals[cat], 2))
        ws.cell(row=r, column=3, value=BUDGETS.get(cat, 0) * 12)

    n = len(sorted_cats)

    # pie chart
    pie = PieChart()
    pie.title = "Expense Distribution by Category"
    pie.width = 16
    pie.height = 12
    cats_ref = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_start + n)
    vals_ref = Reference(ws, min_col=2, min_row=data_start + 1, max_row=data_start + n)
    pie.add_data(vals_ref)
    pie.set_categories(cats_ref)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    ws.add_chart(pie, "A12")

    # bar chart: budget vs actual
    bar = BarChart()
    bar.type = "bar"
    bar.title = "Budget vs Actual by Category"
    bar.y_axis.title = "Category"
    bar.x_axis.title = "Amount ($)"
    bar.y_axis.delete = False
    bar.x_axis.delete = False
    bar.x_axis.numFmt = '$#,##0'
    bar.width = 18
    bar.height = 12

    actual_ref = Reference(ws, min_col=2, min_row=data_start, max_row=data_start + n)
    budget_ref = Reference(ws, min_col=3, min_row=data_start, max_row=data_start + n)
    bar.add_data(actual_ref, titles_from_data=True)
    bar.add_data(budget_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.series[0].graphicalProperties.solidFill = RED
    bar.series[1].graphicalProperties.solidFill = MED_BLUE
    ws.add_chart(bar, "A30")

    return ws


# ── SHEET 5: Dashboard ─────────────────────────────────────────
def build_dashboard(wb, txns):
    ws = wb.create_sheet("Dashboard")

    # compute aggregates
    monthly_income = [0] * 12
    monthly_expense = [0] * 12
    cat_totals = {}
    for t in txns:
        m = t["date"].month - 1
        if t["type"] == "Income":
            monthly_income[m] += t["amount"]
        else:
            monthly_expense[m] += t["amount"]
            cat_totals[t["category"]] = cat_totals.get(t["category"], 0) + t["amount"]

    total_inc = sum(monthly_income)
    total_exp = sum(monthly_expense)
    net_sav = total_inc - total_exp
    sav_rate = net_sav / total_inc if total_inc else 0
    best_month_idx = max(range(12), key=lambda i: monthly_income[i] - monthly_expense[i])
    best_month = MONTHS[best_month_idx]
    most_exp_cat = max(cat_totals, key=cat_totals.get)
    total_budget = sum(BUDGETS.values()) * 12
    overall_variance = total_budget - total_exp
    budget_health = "On Track" if overall_variance >= 0 else "Over Budget"
    sorted_cats = sorted(cat_totals.keys(), key=lambda c: cat_totals[c], reverse=True)

    # title
    ws.merge_cells("A1:L2")
    title_cell = ws.cell(row=1, column=1, value="Financial Budget Tracker | 2024")
    title_cell.font = Font(name="Calibri", bold=True, size=22, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # KPI cards row 4
    kpis = [
        ("Total Income",          f"${total_inc:,.2f}",   GREEN),
        ("Total Expenses",        f"${total_exp:,.2f}",   RED),
        ("Net Savings",           f"${net_sav:,.2f}",     MED_BLUE),
        ("Savings Rate",          f"{sav_rate:.1%}",       BLUE_KPI),
        ("Best Month",            best_month,              GREEN),
        ("Most Expensive Cat.",    most_exp_cat,            RED),
    ]

    for i, (label, value, color) in enumerate(kpis):
        col = 1 + i * 2
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        lbl_cell = ws.cell(row=4, column=col, value=label)
        lbl_cell.font = Font(bold=True, color=WHITE, size=10)
        lbl_cell.fill = PatternFill("solid", fgColor=color)
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
        # fill merged area
        ws.cell(row=4, column=col + 1).fill = PatternFill("solid", fgColor=color)

        val_cell = ws.cell(row=5, column=col, value=value)
        val_cell.font = Font(bold=True, size=14, color=DARK_BLUE)
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws.cell(row=5, column=col + 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    # Budget health
    ws.merge_cells("A7:D7")
    health_cell = ws.cell(row=7, column=1, value=f"Budget Health: {budget_health}")
    health_color = GREEN if overall_variance >= 0 else RED
    health_cell.font = Font(bold=True, size=14, color=WHITE)
    health_cell.fill = PatternFill("solid", fgColor=health_color)
    health_cell.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(2, 5):
        ws.cell(row=7, column=c).fill = PatternFill("solid", fgColor=health_color)

    # chart data at row 200+
    data_start = 200
    ws.cell(row=data_start, column=1, value="Month")
    ws.cell(row=data_start, column=2, value="Income")
    ws.cell(row=data_start, column=3, value="Expenses")
    ws.cell(row=data_start, column=4, value="Net Savings")
    for i in range(12):
        r = data_start + 1 + i
        ws.cell(row=r, column=1, value=MONTHS[i])
        ws.cell(row=r, column=2, value=round(monthly_income[i], 2))
        ws.cell(row=r, column=3, value=round(monthly_expense[i], 2))
        ws.cell(row=r, column=4, value=round(monthly_income[i] - monthly_expense[i], 2))

    # category data
    ws.cell(row=data_start, column=6, value="Category")
    ws.cell(row=data_start, column=7, value="Amount")
    for i, cat in enumerate(sorted_cats):
        ws.cell(row=data_start + 1 + i, column=6, value=cat)
        ws.cell(row=data_start + 1 + i, column=7, value=round(cat_totals[cat], 2))

    cats_ref = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_start + 12)
    n_cats = len(sorted_cats)

    # line chart
    line = LineChart()
    line.title = "Income vs Expenses Trend"
    line.y_axis.delete = False
    line.x_axis.delete = False
    line.y_axis.numFmt = '$#,##0'
    line.width = 18
    line.height = 11
    inc_ref = Reference(ws, min_col=2, min_row=data_start, max_row=data_start + 12)
    exp_ref = Reference(ws, min_col=3, min_row=data_start, max_row=data_start + 12)
    line.add_data(inc_ref, titles_from_data=True)
    line.add_data(exp_ref, titles_from_data=True)
    line.set_categories(cats_ref)
    line.series[0].graphicalProperties.line.solidFill = GREEN
    line.series[1].graphicalProperties.line.solidFill = RED
    ws.add_chart(line, "A9")

    # pie chart
    pie = PieChart()
    pie.title = "Expense Distribution"
    pie.width = 14
    pie.height = 11
    pie_cats = Reference(ws, min_col=6, min_row=data_start + 1, max_row=data_start + n_cats)
    pie_vals = Reference(ws, min_col=7, min_row=data_start + 1, max_row=data_start + n_cats)
    pie.add_data(pie_vals)
    pie.set_categories(pie_cats)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws.add_chart(pie, "J9")

    # bar chart: net savings
    bar = BarChart()
    bar.type = "col"
    bar.title = "Monthly Net Savings"
    bar.y_axis.delete = False
    bar.x_axis.delete = False
    bar.y_axis.numFmt = '$#,##0'
    bar.width = 18
    bar.height = 11
    net_ref = Reference(ws, min_col=4, min_row=data_start, max_row=data_start + 12)
    bar.add_data(net_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.series[0].graphicalProperties.solidFill = MED_BLUE
    ws.add_chart(bar, "A26")

    # top 5 expense categories table
    ws.cell(row=26, column=10, value="Top 5 Expense Categories").font = Font(bold=True, size=12, color=DARK_BLUE)
    for c, h in enumerate(["#", "Category", "Total"], 10):
        cell = ws.cell(row=27, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for i, cat in enumerate(sorted_cats[:5]):
        r = 28 + i
        ws.cell(row=r, column=10, value=i + 1).border = thin_border
        ws.cell(row=r, column=11, value=cat).border = thin_border
        c = ws.cell(row=r, column=12, value=round(cat_totals[cat], 2))
        c.number_format = currency_fmt
        c.border = thin_border
        if i % 2 == 1:
            for cc in range(10, 13):
                ws.cell(row=r, column=cc).fill = alt_fill

    # key insights
    monthly_net = [monthly_income[i] - monthly_expense[i] for i in range(12)]
    worst_month = MONTHS[min(range(12), key=lambda i: monthly_net[i])]
    avg_savings = net_sav / 12

    insights = [
        f"Average monthly savings: ${avg_savings:,.2f}",
        f"Best performing month: {best_month} with highest net savings",
        f"Lowest performing month: {worst_month}",
        f"Largest expense category: {most_exp_cat} (${cat_totals[most_exp_cat]:,.2f})",
    ]

    ws.cell(row=43, column=1, value="Key Insights").font = Font(bold=True, size=12, color=DARK_BLUE)
    for i, insight in enumerate(insights):
        r = 44 + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        cell = ws.cell(row=r, column=1, value=f"  \u2022  {insight}")
        cell.font = Font(size=11, color=DARK_GRAY)
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        for cc in range(2, 9):
            ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    # column widths
    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 15

    return ws


# ── SHEET 6: Yearly Report ─────────────────────────────────────
def build_yearly_report(wb, txns):
    ws = wb.create_sheet("Yearly Report")

    monthly_income = [0] * 12
    monthly_expense = [0] * 12
    cat_totals = {}
    for t in txns:
        m = t["date"].month - 1
        if t["type"] == "Income":
            monthly_income[m] += t["amount"]
        else:
            monthly_expense[m] += t["amount"]
            cat_totals[t["category"]] = cat_totals.get(t["category"], 0) + t["amount"]

    total_inc = sum(monthly_income)
    total_exp = sum(monthly_expense)
    net_sav = total_inc - total_exp

    # title
    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1, value="Annual Financial Report — 2024")
    title.font = Font(bold=True, size=16, color=WHITE)
    title.fill = PatternFill("solid", fgColor=DARK_BLUE)
    title.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(2, 7):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=DARK_BLUE)

    # annual summary
    ws.cell(row=3, column=1, value="Annual Summary").font = Font(bold=True, size=13, color=DARK_BLUE)
    summary_data = [
        ("Total Income",          total_inc),
        ("Total Expenses",        total_exp),
        ("Net Savings",           net_sav),
        ("Savings Rate",          net_sav / total_inc if total_inc else 0),
        ("Avg Monthly Income",    total_inc / 12),
        ("Avg Monthly Expenses",  total_exp / 12),
        ("Avg Monthly Savings",   net_sav / 12),
    ]
    for i, (label, val) in enumerate(summary_data):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r, column=1).border = thin_border
        cell = ws.cell(row=r, column=2, value=round(val, 4) if "Rate" in label else round(val, 2))
        cell.border = thin_border
        cell.number_format = pct_fmt if "Rate" in label else currency_fmt

    # quarterly breakdown
    ws.cell(row=13, column=1, value="Quarterly Breakdown").font = Font(bold=True, size=13, color=DARK_BLUE)
    q_headers = ["Quarter", "Income", "Expenses", "Net Savings", "Savings Rate"]
    for c, h in enumerate(q_headers, 1):
        ws.cell(row=14, column=c, value=h)
    style_header(ws, 14, len(q_headers))

    for q in range(4):
        r = 15 + q
        months_in_q = range(q * 3, q * 3 + 3)
        q_inc = sum(monthly_income[m] for m in months_in_q)
        q_exp = sum(monthly_expense[m] for m in months_in_q)
        q_net = q_inc - q_exp
        q_rate = q_net / q_inc if q_inc else 0

        ws.cell(row=r, column=1, value=f"Q{q+1}")
        ws.cell(row=r, column=2, value=round(q_inc, 2)).number_format = currency_fmt
        ws.cell(row=r, column=3, value=round(q_exp, 2)).number_format = currency_fmt
        ws.cell(row=r, column=4, value=round(q_net, 2)).number_format = currency_fmt
        ws.cell(row=r, column=5, value=round(q_rate, 4)).number_format = pct_fmt

    style_data_area(ws, 15, 18, 5, currency_cols=[2, 3, 4], pct_cols=[5])

    # year-end projections (already December, so use actual)
    ws.cell(row=21, column=1, value="Year-End Projection").font = Font(bold=True, size=13, color=DARK_BLUE)
    ws.cell(row=22, column=1, value="Projected Annual Income").font = Font(bold=True)
    ws.cell(row=22, column=2, value=round(total_inc, 2)).number_format = currency_fmt
    ws.cell(row=23, column=1, value="Projected Annual Expenses").font = Font(bold=True)
    ws.cell(row=23, column=2, value=round(total_exp, 2)).number_format = currency_fmt
    ws.cell(row=24, column=1, value="Projected Net Savings").font = Font(bold=True)
    ws.cell(row=24, column=2, value=round(net_sav, 2)).number_format = currency_fmt
    for r in range(22, 25):
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2).border = thin_border

    # category comparison
    ws.cell(row=27, column=1, value="Category Budget Comparison").font = Font(bold=True, size=13, color=DARK_BLUE)
    cat_headers = ["Category", "Annual Budget", "Actual Spent", "Variance", "Status"]
    for c, h in enumerate(cat_headers, 1):
        ws.cell(row=28, column=c, value=h)
    style_header(ws, 28, len(cat_headers))

    sorted_cats = sorted(cat_totals.keys(), key=lambda c: cat_totals[c], reverse=True)
    for i, cat in enumerate(sorted_cats):
        r = 29 + i
        budget = BUDGETS.get(cat, 0) * 12
        actual = round(cat_totals[cat], 2)
        variance = budget - actual
        status = "Under Budget" if variance >= 0 else "Over Budget"

        ws.cell(row=r, column=1, value=cat).font = Font(bold=True)
        ws.cell(row=r, column=2, value=budget).number_format = currency_fmt
        ws.cell(row=r, column=3, value=actual).number_format = currency_fmt
        ws.cell(row=r, column=4, value=round(variance, 2)).number_format = currency_fmt
        status_cell = ws.cell(row=r, column=5, value=status)
        if variance >= 0:
            status_cell.font = Font(color=GREEN, bold=True)
            ws.cell(row=r, column=4).font = Font(color=GREEN)
        else:
            status_cell.font = Font(color=RED, bold=True)
            ws.cell(row=r, column=4).font = Font(color=RED)

    last_cat_row = 28 + len(sorted_cats)
    style_data_area(ws, 29, last_cat_row, 5, currency_cols=[2, 3, 4])

    # recommendations
    rec_start = last_cat_row + 3
    ws.cell(row=rec_start, column=1, value="Recommendations").font = Font(bold=True, size=13, color=DARK_BLUE)
    ws.merge_cells(start_row=rec_start, start_column=1, end_row=rec_start, end_column=6)

    over_budget_cats = [c for c in sorted_cats if BUDGETS.get(c, 0) * 12 < cat_totals[c]]
    recommendations = [
        f"Overall savings rate of {net_sav/total_inc:.1%} — {'excellent' if net_sav/total_inc > 0.15 else 'needs improvement'}.",
        f"Consider increasing the budget for: {', '.join(over_budget_cats[:3]) if over_budget_cats else 'none needed — all categories are within budget'}.",
        "Review recurring subscriptions quarterly to eliminate unnecessary costs.",
        "Aim to build 3-6 months of expenses in your emergency fund.",
        f"Top spending area is {sorted_cats[0]} (${cat_totals[sorted_cats[0]]:,.2f}) — look for optimization opportunities.",
        "Consider automating savings transfers on payday for consistency.",
    ]

    for i, rec in enumerate(recommendations):
        r = rec_start + 1 + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        cell = ws.cell(row=r, column=1, value=f"  \u2022  {rec}")
        cell.font = Font(size=11, color=DARK_GRAY)
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        for cc in range(2, 7):
            ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    auto_width(ws, 6, min_width=16, max_width=24)
    return ws


# ── MAIN ────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    txns = generate_transactions()
    print(f"Generated {len(txns)} transactions")

    build_transactions_sheet(wb, txns)
    build_monthly_budget(wb, txns)
    build_monthly_summary(wb, txns)
    build_category_breakdown(wb, txns)
    build_dashboard(wb, txns)
    build_yearly_report(wb, txns)

    out = "budget_tracker.xlsx"
    wb.save(out)
    print(f"Saved {out}")


if __name__ == "__main__":
    main()
