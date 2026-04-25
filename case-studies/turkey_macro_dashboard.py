"""
Turkey Inflation & Unemployment Dashboard (2019-2025)
=====================================================
Generates a professional Excel dashboard with:
- Raw Data sheet: monthly TÜFE (CPI YoY%) and unemployment rate
- Analysis sheet: yearly averages, correlation, summary stats
- Dashboard sheet: charts, KPIs, conditional formatting

Data sources: TÜİK (Turkish Statistical Institute), TCMB (Central Bank)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from datetime import datetime
import calendar

# =============================================================================
# REAL DATA — Source: TÜİK & TCMB official releases
# Monthly CPI (TÜFE) Year-over-Year % change
# Monthly Unemployment Rate %
# =============================================================================

DATA = {
    # (month, year): (tufe_yoy%, unemployment%)
    # 2019 — Inflation cooling from 2018 crisis, unemployment spike
    (1, 2019): (20.35, 14.7), (2, 2019): (19.67, 15.0), (3, 2019): (19.71, 14.1),
    (4, 2019): (19.50, 13.0), (5, 2019): (18.71, 12.8), (6, 2019): (15.72, 13.0),
    (7, 2019): (16.65, 13.9), (8, 2019): (15.01, 13.8), (9, 2019): (9.26, 13.8),
    (10, 2019): (8.55, 13.4), (11, 2019): (10.56, 13.3), (12, 2019): (11.84, 13.7),

    # 2020 — COVID impact, moderate inflation, unemployment jump
    (1, 2020): (12.15, 13.5), (2, 2020): (12.37, 13.6), (3, 2020): (11.86, 13.2),
    (4, 2020): (10.94, 12.8), (5, 2020): (11.39, 12.9), (6, 2020): (12.62, 13.4),
    (7, 2020): (11.76, 13.0), (8, 2020): (11.77, 13.2), (9, 2020): (11.75, 12.7),
    (10, 2020): (11.89, 12.7), (11, 2020): (14.03, 12.9), (12, 2020): (14.60, 13.2),

    # 2021 — Inflation starts climbing, TCMB rate cuts begin
    (1, 2021): (14.97, 13.4), (2, 2021): (15.61, 13.4), (3, 2021): (16.19, 13.1),
    (4, 2021): (17.14, 13.1), (5, 2021): (16.59, 12.8), (6, 2021): (17.53, 12.0),
    (7, 2021): (18.95, 12.0), (8, 2021): (19.25, 11.7), (9, 2021): (19.58, 11.1),
    (10, 2021): (19.89, 11.2), (11, 2021): (21.31, 11.1), (12, 2021): (36.08, 11.2),

    # 2022 — Hyperinflation, lira collapse
    (1, 2022): (48.69, 11.4), (2, 2022): (54.44, 11.4), (3, 2022): (61.14, 11.5),
    (4, 2022): (69.97, 11.3), (5, 2022): (73.50, 10.7), (6, 2022): (78.62, 10.3),
    (7, 2022): (79.60, 10.1), (8, 2022): (80.21, 10.1), (9, 2022): (83.45, 10.1),
    (10, 2022): (85.51, 10.2), (11, 2022): (84.39, 10.3), (12, 2022): (64.27, 10.3),

    # 2023 — Election year, continued high inflation
    (1, 2023): (57.68, 10.4), (2, 2023): (55.18, 10.0), (3, 2023): (50.51, 10.2),
    (4, 2023): (43.68, 10.0), (5, 2023): (39.59, 9.6), (6, 2023): (38.21, 9.6),
    (7, 2023): (47.83, 9.4), (8, 2023): (58.94, 9.2), (9, 2023): (61.53, 8.9),
    (10, 2023): (61.36, 8.6), (11, 2023): (61.98, 8.5), (12, 2023): (64.77, 8.7),

    # 2024 — Tight monetary policy begins, slow disinflation
    (1, 2024): (64.86, 8.8), (2, 2024): (67.07, 8.7), (3, 2024): (68.50, 8.4),
    (4, 2024): (69.80, 8.5), (5, 2024): (75.45, 8.4), (6, 2024): (71.60, 8.5),
    (7, 2024): (61.78, 8.2), (8, 2024): (51.97, 8.3), (9, 2024): (49.38, 8.3),
    (10, 2024): (48.58, 8.2), (11, 2024): (47.09, 8.1), (12, 2024): (44.38, 8.1),

    # 2025 — Disinflation trend continues (latest available data)
    (1, 2025): (42.12, 8.3), (2, 2025): (39.05, 8.5), (3, 2025): (38.10, 8.2),
    (4, 2025): (37.86, 8.0),
}


def create_workbook():
    wb = openpyxl.Workbook()

    # Color scheme
    DARK_BLUE = "1B3A5C"
    MED_BLUE = "2E75B6"
    LIGHT_BLUE = "D6E4F0"
    DARK_GRAY = "404040"
    LIGHT_GRAY = "F2F2F2"
    WHITE = "FFFFFF"
    RED = "C00000"
    GREEN = "548235"

    header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    subheader_fill = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
    subheader_font = Font(name="Calibri", bold=True, color=WHITE, size=10)
    data_font = Font(name="Calibri", size=10, color=DARK_GRAY)
    alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    # =========================================================================
    # SHEET 1: Raw Data (Ham Veri)
    # =========================================================================
    ws_raw = wb.active
    ws_raw.title = "Ham Veri"
    ws_raw.sheet_properties.tabColor = DARK_BLUE

    # Header
    headers = ["Tarih", "Ay", "Yıl", "TÜFE YoY (%)", "İşsizlik (%)"]
    for col, h in enumerate(headers, 1):
        cell = ws_raw.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    sorted_keys = sorted(DATA.keys(), key=lambda x: (x[1], x[0]))
    for i, (month, year) in enumerate(sorted_keys):
        row = i + 2
        tufe, unemp = DATA[(month, year)]
        date_str = f"{calendar.month_abbr[month]}-{year}"

        ws_raw.cell(row=row, column=1, value=date_str).font = data_font
        ws_raw.cell(row=row, column=2, value=month).font = data_font
        ws_raw.cell(row=row, column=3, value=year).font = data_font
        ws_raw.cell(row=row, column=4, value=tufe).font = data_font
        ws_raw.cell(row=row, column=5, value=unemp).font = data_font

        for col in range(1, 6):
            c = ws_raw.cell(row=row, column=col)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center")
            if i % 2 == 1:
                c.fill = alt_fill

    # Number format
    total_rows = len(sorted_keys)
    for row in range(2, total_rows + 2):
        ws_raw.cell(row=row, column=4).number_format = '0.00"%"'
        ws_raw.cell(row=row, column=5).number_format = '0.0"%"'

    # Column widths
    ws_raw.column_dimensions["A"].width = 12
    ws_raw.column_dimensions["B"].width = 6
    ws_raw.column_dimensions["C"].width = 8
    ws_raw.column_dimensions["D"].width = 16
    ws_raw.column_dimensions["E"].width = 14

    # Freeze panes
    ws_raw.freeze_panes = "A2"

    # =========================================================================
    # SHEET 2: Analysis (Analiz)
    # =========================================================================
    ws_analysis = wb.create_sheet("Analiz")
    ws_analysis.sheet_properties.tabColor = MED_BLUE

    # --- Yearly averages ---
    ws_analysis.cell(row=1, column=1, value="Yıllık Ortalamalar").font = Font(
        name="Calibri", bold=True, size=14, color=DARK_BLUE
    )
    ws_analysis.merge_cells("A1:E1")

    year_headers = ["Yıl", "Ort. TÜFE (%)", "Min TÜFE (%)", "Max TÜFE (%)",
                    "Ort. İşsizlik (%)", "Min İşsizlik (%)", "Max İşsizlik (%)"]
    for col, h in enumerate(year_headers, 1):
        cell = ws_analysis.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    years = sorted(set(k[1] for k in DATA.keys()))
    for i, year in enumerate(years):
        row = i + 4
        year_data = {k: v for k, v in DATA.items() if k[1] == year}
        tufe_vals = [v[0] for v in year_data.values()]
        unemp_vals = [v[1] for v in year_data.values()]

        values = [
            year,
            round(sum(tufe_vals) / len(tufe_vals), 2),
            min(tufe_vals),
            max(tufe_vals),
            round(sum(unemp_vals) / len(unemp_vals), 2),
            min(unemp_vals),
            max(unemp_vals),
        ]
        for col, val in enumerate(values, 1):
            cell = ws_analysis.cell(row=row, column=col, value=val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            if i % 2 == 1:
                cell.fill = alt_fill

    # --- Correlation ---
    corr_row = 4 + len(years) + 2
    ws_analysis.cell(row=corr_row, column=1, value="Korelasyon Analizi").font = Font(
        name="Calibri", bold=True, size=14, color=DARK_BLUE
    )
    ws_analysis.merge_cells(f"A{corr_row}:E{corr_row}")

    # CORREL formula referencing Ham Veri sheet
    last_data_row = total_rows + 1
    ws_analysis.cell(row=corr_row + 2, column=1, value="TÜFE ↔ İşsizlik Korelasyonu:").font = Font(
        name="Calibri", bold=True, size=11, color=DARK_BLUE
    )
    corr_cell = ws_analysis.cell(
        row=corr_row + 2, column=3,
        value=f"=CORREL('Ham Veri'!D2:D{last_data_row},'Ham Veri'!E2:E{last_data_row})"
    )
    corr_cell.font = Font(name="Calibri", bold=True, size=14, color=RED)
    corr_cell.number_format = '0.000'

    ws_analysis.cell(row=corr_row + 4, column=1,
                     value="Yorum: Negatif korelasyon → enflasyon yükselirken işsizlik düşme eğiliminde").font = Font(
        name="Calibri", italic=True, size=10, color=DARK_GRAY
    )
    ws_analysis.merge_cells(f"A{corr_row + 4}:G{corr_row + 4}")

    # --- Summary stats ---
    stats_row = corr_row + 7
    ws_analysis.cell(row=stats_row, column=1, value="Özet İstatistikler").font = Font(
        name="Calibri", bold=True, size=14, color=DARK_BLUE
    )
    ws_analysis.merge_cells(f"A{stats_row}:E{stats_row}")

    all_tufe = [v[0] for v in DATA.values()]
    all_unemp = [v[1] for v in DATA.values()]

    stats = [
        ("Toplam Gözlem Sayısı", len(DATA), "ay"),
        ("Dönem", "Oca 2019 – Nis 2025", ""),
        ("", "", ""),
        ("TÜFE Ortalaması", round(sum(all_tufe) / len(all_tufe), 2), "%"),
        ("TÜFE Minimum", f"{min(all_tufe):.2f} (Eki 2019)", "%"),
        ("TÜFE Maksimum", f"{max(all_tufe):.2f} (Eki 2022)", "%"),
        ("", "", ""),
        ("İşsizlik Ortalaması", round(sum(all_unemp) / len(all_unemp), 2), "%"),
        ("İşsizlik Minimum", f"{min(all_unemp):.1f} (Nis 2025)", "%"),
        ("İşsizlik Maksimum", f"{max(all_unemp):.1f} (Şub 2019)", "%"),
    ]
    for j, (label, value, unit) in enumerate(stats):
        r = stats_row + 2 + j
        ws_analysis.cell(row=r, column=1, value=label).font = Font(
            name="Calibri", bold=True, size=10, color=DARK_GRAY
        )
        ws_analysis.cell(row=r, column=3, value=value).font = Font(
            name="Calibri", size=10, color=DARK_BLUE
        )
        ws_analysis.cell(row=r, column=4, value=unit).font = data_font

    # Column widths
    for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
        ws_analysis.column_dimensions[col_letter].width = 18

    # =========================================================================
    # SHEET 3: Dashboard
    # =========================================================================
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash.sheet_properties.tabColor = "1B3A5C"

    # --- Title ---
    ws_dash.merge_cells("A1:N2")
    title_cell = ws_dash.cell(row=1, column=1,
                              value="Turkey Inflation & Unemployment Analysis  |  2019 – 2025")
    title_cell.font = Font(name="Calibri", bold=True, size=22, color=WHITE)
    title_cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Subtitle
    ws_dash.merge_cells("A3:N3")
    sub_cell = ws_dash.cell(row=3, column=1,
                            value="Data Source: TÜİK (Turkish Statistical Institute) & TCMB (Central Bank of Turkey)")
    sub_cell.font = Font(name="Calibri", size=10, color=WHITE, italic=True)
    sub_cell.fill = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- KPI Cards ---
    kpi_row = 5
    kpis = [
        ("Peak Inflation", f"{max(all_tufe):.1f}%", "Oct 2022", RED),
        ("Latest Inflation", f"{all_tufe[-1] if sorted_keys else 0:.1f}%", "Apr 2025", MED_BLUE),
        ("Avg Unemployment", f"{sum(all_unemp)/len(all_unemp):.1f}%", "2019-2025", DARK_BLUE),
        ("Latest Unemployment", f"{all_unemp[-1] if sorted_keys else 0:.1f}%", "Apr 2025", GREEN),
    ]

    for i, (title, value, subtitle, color) in enumerate(kpis):
        col_start = 1 + i * 4
        col_end = col_start + 2

        # Card background
        card_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for r in range(kpi_row, kpi_row + 3):
            for c in range(col_start, col_end + 1):
                ws_dash.cell(row=r, column=c).fill = card_fill

        # Title
        ws_dash.merge_cells(start_row=kpi_row, start_column=col_start,
                           end_row=kpi_row, end_column=col_end)
        ws_dash.cell(row=kpi_row, column=col_start, value=title).font = Font(
            name="Calibri", size=9, color=WHITE
        )
        ws_dash.cell(row=kpi_row, column=col_start).alignment = Alignment(
            horizontal="center", vertical="center"
        )

        # Value
        ws_dash.merge_cells(start_row=kpi_row + 1, start_column=col_start,
                           end_row=kpi_row + 1, end_column=col_end)
        ws_dash.cell(row=kpi_row + 1, column=col_start, value=value).font = Font(
            name="Calibri", bold=True, size=20, color=WHITE
        )
        ws_dash.cell(row=kpi_row + 1, column=col_start).alignment = Alignment(
            horizontal="center", vertical="center"
        )

        # Subtitle
        ws_dash.merge_cells(start_row=kpi_row + 2, start_column=col_start,
                           end_row=kpi_row + 2, end_column=col_end)
        ws_dash.cell(row=kpi_row + 2, column=col_start, value=subtitle).font = Font(
            name="Calibri", size=8, color=WHITE, italic=True
        )
        ws_dash.cell(row=kpi_row + 2, column=col_start).alignment = Alignment(
            horizontal="center", vertical="center"
        )

    # --- Chart data in a dedicated area on Dashboard (row 100+, invisible to user) ---
    last_data_row_raw = total_rows + 1
    cdata_row = 100  # far below visible area

    # Write chart-specific data: col A = label, col B = TÜFE, col C = unemployment
    ws_dash.cell(row=cdata_row, column=1, value="Month")
    ws_dash.cell(row=cdata_row, column=2, value="CPI Inflation (%)")
    ws_dash.cell(row=cdata_row, column=3, value="Unemployment (%)")

    for i, (month, year) in enumerate(sorted_keys):
        r = cdata_row + 1 + i
        tufe, unemp = DATA[(month, year)]
        ws_dash.cell(row=r, column=1, value=f"{calendar.month_abbr[month]} {year}")
        ws_dash.cell(row=r, column=2, value=tufe)
        ws_dash.cell(row=r, column=3, value=unemp)

    cdata_end = cdata_row + total_rows  # last data row

    # --- CHART 1: Inflation Line Chart ---
    chart1 = LineChart()
    chart1.title = "Monthly CPI Inflation Rate (YoY %)"
    chart1.style = 10
    chart1.width = 28
    chart1.height = 14

    chart1.y_axis.title = "Inflation Rate (%)"
    chart1.y_axis.numFmt = '0.0"%"'
    chart1.y_axis.scaling.min = 0
    chart1.y_axis.scaling.max = 90
    chart1.y_axis.delete = False
    chart1.y_axis.majorGridlines = None

    # Categories = month labels (col A)
    cats1 = Reference(ws_dash, min_col=1, min_row=cdata_row + 1, max_row=cdata_end)
    # Data = TÜFE values (col B), with header
    data1 = Reference(ws_dash, min_col=2, min_row=cdata_row, max_row=cdata_end)
    chart1.add_data(data1, titles_from_data=True, from_rows=False)
    chart1.set_categories(cats1)

    # X-axis: show every 6th label, rotated for readability
    chart1.x_axis.tickLblSkip = 6
    chart1.x_axis.tickMarkSkip = 6
    chart1.x_axis.txPr = None  # reset any text properties
    chart1.x_axis.delete = False

    s1 = chart1.series[0]
    s1.graphicalProperties.line.solidFill = RED
    s1.graphicalProperties.line.width = 22000
    s1.smooth = False  # no smoothing for accuracy

    ws_dash.add_chart(chart1, "A10")

    # --- CHART 2: Unemployment Line Chart ---
    chart2 = LineChart()
    chart2.title = "Monthly Unemployment Rate (%)"
    chart2.style = 10
    chart2.width = 28
    chart2.height = 14

    chart2.y_axis.title = "Unemployment Rate (%)"
    chart2.y_axis.numFmt = '0.0"%"'
    chart2.y_axis.scaling.min = 7
    chart2.y_axis.scaling.max = 16
    chart2.y_axis.delete = False
    chart2.y_axis.majorGridlines = None

    cats2 = Reference(ws_dash, min_col=1, min_row=cdata_row + 1, max_row=cdata_end)
    data2 = Reference(ws_dash, min_col=3, min_row=cdata_row, max_row=cdata_end)
    chart2.add_data(data2, titles_from_data=True, from_rows=False)
    chart2.set_categories(cats2)

    chart2.x_axis.tickLblSkip = 6
    chart2.x_axis.tickMarkSkip = 6
    chart2.x_axis.delete = False

    s2 = chart2.series[0]
    s2.graphicalProperties.line.solidFill = MED_BLUE
    s2.graphicalProperties.line.width = 22000
    s2.smooth = False

    ws_dash.add_chart(chart2, "A27")

    # --- Yearly averages table on dashboard ---
    table_start_col = 10  # Column J
    table_row = 10

    ws_dash.merge_cells(start_row=table_row, start_column=table_start_col,
                       end_row=table_row, end_column=table_start_col + 3)
    ws_dash.cell(row=table_row, column=table_start_col,
                 value="Yearly Averages").font = Font(
        name="Calibri", bold=True, size=13, color=DARK_BLUE
    )

    t_headers = ["Year", "Avg CPI (%)", "Avg Unemp (%)", "CORREL"]
    for j, h in enumerate(t_headers):
        cell = ws_dash.cell(row=table_row + 1, column=table_start_col + j, value=h)
        cell.fill = header_fill
        cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, year in enumerate(years):
        r = table_row + 2 + i
        year_data = {k: v for k, v in DATA.items() if k[1] == year}
        tufe_vals = [v[0] for v in year_data.values()]
        unemp_vals = [v[1] for v in year_data.values()]

        avg_tufe = round(sum(tufe_vals) / len(tufe_vals), 2)
        avg_unemp = round(sum(unemp_vals) / len(unemp_vals), 2)

        # Compute correlation for the year
        n = len(tufe_vals)
        if n > 1:
            mean_t = sum(tufe_vals) / n
            mean_u = sum(unemp_vals) / n
            cov = sum((t - mean_t) * (u - mean_u) for t, u in zip(tufe_vals, unemp_vals))
            std_t = (sum((t - mean_t) ** 2 for t in tufe_vals)) ** 0.5
            std_u = (sum((u - mean_u) ** 2 for u in unemp_vals)) ** 0.5
            corr = round(cov / (std_t * std_u), 3) if std_t * std_u > 0 else 0
        else:
            corr = "N/A"

        vals = [year, avg_tufe, avg_unemp, corr]
        for j, val in enumerate(vals):
            cell = ws_dash.cell(row=r, column=table_start_col + j, value=val)
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            if i % 2 == 1:
                cell.fill = alt_fill

    # Conditional formatting — highlight high inflation years
    for i, year in enumerate(years):
        r = table_row + 2 + i
        cell_ref = f"{get_column_letter(table_start_col + 1)}{r}"
        year_data = {k: v for k, v in DATA.items() if k[1] == year}
        avg_tufe = sum(v[0] for v in year_data.values()) / len(year_data)
        if avg_tufe > 50:
            ws_dash.cell(row=r, column=table_start_col + 1).fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )
            ws_dash.cell(row=r, column=table_start_col + 1).font = Font(
                name="Calibri", size=10, color=RED, bold=True
            )

    # Conditional formatting on raw data — color scale for TÜFE
    ws_raw.conditional_formatting.add(
        f"D2:D{total_rows + 1}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        )
    )

    # Conditional formatting on raw data — color scale for unemployment
    ws_raw.conditional_formatting.add(
        f"E2:E{total_rows + 1}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        )
    )

    # Overall correlation on dashboard
    overall_corr_row = table_row + 2 + len(years) + 1
    ws_dash.cell(row=overall_corr_row, column=table_start_col,
                 value="Overall CORREL:").font = Font(
        name="Calibri", bold=True, size=11, color=DARK_BLUE
    )
    ws_dash.cell(row=overall_corr_row, column=table_start_col + 1,
                 value=f"=CORREL(B{cdata_row+1}:B{cdata_end},C{cdata_row+1}:C{cdata_end})"
    ).font = Font(name="Calibri", bold=True, size=14, color=RED)
    ws_dash.cell(row=overall_corr_row, column=table_start_col + 1).number_format = '0.000'

    # Key insight box
    insight_row = overall_corr_row + 2
    ws_dash.merge_cells(start_row=insight_row, start_column=table_start_col,
                       end_row=insight_row + 3, end_column=table_start_col + 3)
    insight = ws_dash.cell(row=insight_row, column=table_start_col)
    insight.value = (
        "KEY INSIGHT\n"
        "Turkey experienced severe inflation (85%+ peak)\n"
        "while unemployment fell — a classic Phillips\n"
        "Curve pattern amplified by currency crisis."
    )
    insight.font = Font(name="Calibri", size=10, color=DARK_BLUE)
    insight.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    insight_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    for r in range(insight_row, insight_row + 4):
        for c in range(table_start_col, table_start_col + 4):
            ws_dash.cell(row=r, column=c).fill = insight_fill

    # --- Column widths for dashboard ---
    for c in range(1, 15):
        ws_dash.column_dimensions[get_column_letter(c)].width = 10

    # Row heights for title
    ws_dash.row_dimensions[1].height = 30
    ws_dash.row_dimensions[2].height = 20
    ws_dash.row_dimensions[3].height = 20
    ws_dash.row_dimensions[kpi_row + 1].height = 35

    # Set Dashboard as active sheet
    wb.active = wb.sheetnames.index("Dashboard")

    # Print settings
    ws_dash.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

    return wb


if __name__ == "__main__":
    wb = create_workbook()
    filename = "turkey_macro_dashboard.xlsx"
    wb.save(filename)
    print(f"Dashboard saved: {filename}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Data points: {len(DATA)} months (Jan 2019 – Apr 2025)")
    print(f"Open in Excel to see charts and conditional formatting.")
