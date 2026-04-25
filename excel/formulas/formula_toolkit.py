"""
formula_toolkit.py
------------------
Generates 'formula_toolkit.xlsx' -- an Excel workbook that demonstrates
advanced formula techniques across four themed worksheets.

Run:
    pip install openpyxl
    python formula_toolkit.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from datetime import date

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FORMULA_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, row, max_col, font=SUBHEADER_FONT, fill=SUBHEADER_FILL):
    """Apply header styling to a row of cells."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def write_title(ws, row, title, description, max_col):
    """Write a sheet title block spanning max_col columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=max_col)
    desc_cell = ws.cell(row=row + 1, column=1, value=description)
    desc_cell.font = Font(name="Calibri", italic=True, size=10, color="4472C4")
    desc_cell.alignment = Alignment(horizontal="center", wrap_text=True)


def apply_data_borders(ws, start_row, end_row, start_col, end_col):
    """Apply thin borders to a rectangular data region."""
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def mark_formula(cell, comment_text):
    """Highlight a formula cell and attach an explanatory comment."""
    cell.fill = FORMULA_FILL
    cell.border = THIN_BORDER
    cell.comment = Comment(comment_text, "Formula Toolkit")


def auto_width(ws, min_width=12, max_width=30):
    """Auto-fit column widths based on content length."""
    for col_cells in ws.columns:
        length = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                length = max(length, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = length


# ---------------------------------------------------------------------------
# Sheet 1 -- LOOKUP Examples
# ---------------------------------------------------------------------------

def build_lookup_sheet(wb):
    ws = wb.create_sheet("LOOKUP Examples")

    write_title(ws, 1, "LOOKUP Formula Examples",
                "Demonstrates VLOOKUP, INDEX-MATCH, and XLOOKUP for exact & approximate matching", 8)

    # --- Employee table (columns A-D) ---
    emp_header_row = 4
    headers = ["EmpID", "Name", "Department", "Salary"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=emp_header_row, column=i, value=h)
    style_header_row(ws, emp_header_row, 4)

    employees = [
        [101, "Alice Johnson", "Engineering", 95000],
        [102, "Bob Martinez", "Marketing", 72000],
        [103, "Carol Lee", "Engineering", 105000],
        [104, "David Kim", "Sales", 68000],
        [105, "Eva Chen", "Marketing", 78000],
        [106, "Frank Diaz", "Engineering", 112000],
        [107, "Grace Patel", "Sales", 71000],
        [108, "Hiro Tanaka", "Finance", 88000],
    ]
    for r, row_data in enumerate(employees, emp_header_row + 1):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
    apply_data_borders(ws, emp_header_row, emp_header_row + len(employees), 1, 4)

    # --- Lookup demo area (columns F-H) ---
    ws.cell(row=emp_header_row, column=6, value="Lookup Value")
    ws.cell(row=emp_header_row, column=7, value="Formula")
    ws.cell(row=emp_header_row, column=8, value="Result")
    style_header_row(ws, emp_header_row, 8)
    # Only style cols 6-8 (already styled all 8, which is fine)

    lookups = [
        (103, 'VLOOKUP(F5,A5:D12,3,FALSE)',
         '=VLOOKUP(F5,$A$5:$D$12,3,FALSE)',
         "VLOOKUP -- exact match on EmpID, return Department (col 3)"),
        (106, 'INDEX-MATCH salary',
         '=INDEX($D$5:$D$12,MATCH(F6,$A$5:$A$12,0))',
         "INDEX-MATCH -- more flexible than VLOOKUP; returns Salary for EmpID 106"),
        ("Bob Martinez", 'XLOOKUP by name',
         '=XLOOKUP(F7,$B$5:$B$12,$C$5:$C$12,"Not found")',
         "XLOOKUP -- search by Name, return Department; includes a default if not found"),
        (110, 'XLOOKUP missing ID',
         '=XLOOKUP(F8,$A$5:$A$12,$B$5:$B$12,"Not found")',
         "XLOOKUP -- lookup a non-existent ID to show the default value behavior"),
    ]

    for i, (lookup_val, label, formula, comment) in enumerate(lookups):
        r = emp_header_row + 1 + i
        ws.cell(row=r, column=6, value=lookup_val)
        ws.cell(row=r, column=7, value=label)
        result_cell = ws.cell(row=r, column=8, value=formula)
        mark_formula(result_cell, comment)

    # --- Approximate match table (tax brackets) ---
    tax_row = 15
    ws.merge_cells(start_row=tax_row, start_column=1, end_row=tax_row, end_column=4)
    ws.cell(row=tax_row, column=1, value="Approximate Match -- Tax Bracket Lookup").font = Font(
        bold=True, size=11, color="2F5496"
    )

    tax_headers = ["Min Income", "Tax Rate"]
    for i, h in enumerate(tax_headers, 1):
        ws.cell(row=tax_row + 1, column=i, value=h)
    style_header_row(ws, tax_row + 1, 2)

    brackets = [(0, "10%"), (10000, "12%"), (40000, "22%"),
                (85000, "24%"), (165000, "32%"), (215000, "35%")]
    for r, (inc, rate) in enumerate(brackets, tax_row + 2):
        ws.cell(row=r, column=1, value=inc)
        ws.cell(row=r, column=2, value=rate)
    apply_data_borders(ws, tax_row + 1, tax_row + 1 + len(brackets), 1, 2)

    # Approx-match demos
    ws.cell(row=tax_row + 1, column=4, value="Income")
    ws.cell(row=tax_row + 1, column=5, value="Bracket (approx match)")
    style_header_row(ws, tax_row + 1, 5)

    approx_demos = [
        (55000, '=VLOOKUP(D{r},$A$17:$B$22,2,TRUE)',
         "VLOOKUP approximate match -- TRUE flag finds the largest value <= lookup value"),
    ]
    for i, (val, formula_tmpl, comment) in enumerate(approx_demos):
        r = tax_row + 2 + i
        ws.cell(row=r, column=4, value=val)
        result = ws.cell(row=r, column=5, value=formula_tmpl.format(r=r))
        mark_formula(result, comment)

    auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 2 -- Conditional Formulas
# ---------------------------------------------------------------------------

def build_conditional_sheet(wb):
    ws = wb.create_sheet("Conditional Formulas")

    write_title(ws, 1, "Conditional Aggregation Formulas",
                "SUMIFS, COUNTIFS, AVERAGEIFS with multiple criteria on sales data", 9)

    # --- Sales data ---
    data_row = 4
    headers = ["Date", "Region", "Product", "Rep", "Units", "Revenue"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=data_row, column=i, value=h)
    style_header_row(ws, data_row, 6)

    sales = [
        [date(2025, 1, 5),  "East",  "Widget A", "Alice", 120, 3600],
        [date(2025, 1, 12), "West",  "Widget B", "Bob",    85, 4250],
        [date(2025, 1, 20), "East",  "Widget A", "Alice",  95, 2850],
        [date(2025, 2, 3),  "North", "Widget C", "Carol", 200, 8000],
        [date(2025, 2, 14), "East",  "Widget B", "Alice", 150, 7500],
        [date(2025, 2, 28), "West",  "Widget A", "Bob",    60, 1800],
        [date(2025, 3, 7),  "North", "Widget C", "Carol", 175, 7000],
        [date(2025, 3, 15), "East",  "Widget A", "David",  90, 2700],
        [date(2025, 3, 22), "West",  "Widget B", "Bob",   110, 5500],
        [date(2025, 3, 30), "North", "Widget A", "Carol", 130, 3900],
    ]
    for r, row_data in enumerate(sales, data_row + 1):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 1:
                cell.number_format = "YYYY-MM-DD"
            elif c == 6:
                cell.number_format = "#,##0"
    apply_data_borders(ws, data_row, data_row + len(sales), 1, 6)

    # --- Formula demo area ---
    f_row = 4
    ws.cell(row=f_row, column=8, value="Formula Description")
    ws.cell(row=f_row, column=9, value="Result")
    style_header_row(ws, f_row, 9)
    # Only want cols 8-9 styled; re-style just those
    for c in range(8, 10):
        ws.cell(row=f_row, column=c).font = SUBHEADER_FONT
        ws.cell(row=f_row, column=c).fill = SUBHEADER_FILL

    formulas = [
        ("SUMIFS: Total revenue, East region",
         '=SUMIFS($F$5:$F$14,$B$5:$B$14,"East")',
         "Sum Revenue where Region = East"),
        ("SUMIFS: Revenue, East + Widget A",
         '=SUMIFS($F$5:$F$14,$B$5:$B$14,"East",$C$5:$C$14,"Widget A")',
         "Sum Revenue where Region=East AND Product=Widget A"),
        ("COUNTIFS: Orders by Alice",
         '=COUNTIFS($D$5:$D$14,"Alice")',
         "Count rows where Rep = Alice"),
        ("COUNTIFS: East orders >= 100 units",
         '=COUNTIFS($B$5:$B$14,"East",$E$5:$E$14,">="&100)',
         "Count East-region orders with 100+ units"),
        ("AVERAGEIFS: Avg revenue, West",
         '=AVERAGEIFS($F$5:$F$14,$B$5:$B$14,"West")',
         "Average Revenue for West region"),
        ("AVERAGEIFS: Avg units, Widget C, North",
         '=AVERAGEIFS($E$5:$E$14,$C$5:$C$14,"Widget C",$B$5:$B$14,"North")',
         "Average Units for Widget C in North region"),
        ("SUMIFS: Revenue in Q1 date range",
         '=SUMIFS($F$5:$F$14,$A$5:$A$14,">="&DATE(2025,1,1),$A$5:$A$14,"<="&DATE(2025,1,31))',
         "Sum Revenue for January 2025 only (date-range criteria)"),
    ]

    for i, (desc, formula, comment) in enumerate(formulas):
        r = f_row + 1 + i
        ws.cell(row=r, column=8, value=desc)
        result_cell = ws.cell(row=r, column=9, value=formula)
        mark_formula(result_cell, comment)

    auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 3 -- Text Functions
# ---------------------------------------------------------------------------

def build_text_sheet(wb):
    ws = wb.create_sheet("Text Functions")

    write_title(ws, 1, "Text Function Examples",
                "Data cleanup with CONCATENATE, LEFT, RIGHT, MID, TRIM, SUBSTITUTE and more", 7)

    # --- Messy data ---
    data_row = 4
    headers = ["Raw Data", "Issue", "Formula Description", "Cleaned (Formula)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=data_row, column=i, value=h)
    style_header_row(ws, data_row, 4)

    rows = [
        ("  John   Smith  ", "Extra spaces",
         "TRIM -- remove leading/trailing/extra spaces",
         "=TRIM(A5)",
         "TRIM removes all unnecessary whitespace"),
        ("jane.doe@example.com", "Extract username",
         "LEFT + FIND -- extract text before '@'",
         '=LEFT(A6,FIND("@",A6)-1)',
         "LEFT returns characters from the start; FIND locates the @ position"),
        ("INV-2025-00412", "Extract year",
         "MID -- pull 4 chars starting at position 5",
         "=MID(A7,5,4)",
         "MID(text, start, length) extracts a substring"),
        ("SKU-8842-BLK", "Last 3 chars (color code)",
         "RIGHT -- get last 3 characters",
         "=RIGHT(A8,3)",
         "RIGHT returns characters from the end of a string"),
        ("New York, NY 10001", "Replace comma with ' -'",
         "SUBSTITUTE -- swap specific text",
         '=SUBSTITUTE(A9,","," -")',
         "SUBSTITUTE replaces every occurrence of old text with new text"),
        ("Alice", "Build full greeting",
         "CONCATENATE / & operator",
         '="Hello, "&A10&"! Welcome aboard."',
         "String concatenation using the & operator"),
        ("product_name_final_v2", "Replace underscores with spaces",
         "SUBSTITUTE -- underscores to spaces",
         '=SUBSTITUTE(A11,"_"," ")',
         "SUBSTITUTE is great for batch character replacement"),
        ("2025/04/15", "Reformat date string",
         "MID+LEFT+RIGHT to rearrange",
         '=MID(A12,6,2)&"-"&RIGHT(A12,2)&"-"&LEFT(A12,4)',
         "Combine MID, LEFT, RIGHT to restructure a date string to MM-DD-YYYY"),
    ]

    for i, (raw, issue, desc, formula, comment) in enumerate(rows):
        r = data_row + 1 + i
        ws.cell(row=r, column=1, value=raw)
        ws.cell(row=r, column=2, value=issue)
        ws.cell(row=r, column=3, value=desc)
        result_cell = ws.cell(row=r, column=4, value=formula)
        mark_formula(result_cell, comment)
    apply_data_borders(ws, data_row, data_row + len(rows), 1, 4)

    # --- Bonus: TEXTJOIN / UPPER / LOWER / PROPER ---
    bonus_row = data_row + len(rows) + 3
    ws.merge_cells(start_row=bonus_row, start_column=1, end_row=bonus_row, end_column=4)
    ws.cell(row=bonus_row, column=1,
            value="Bonus -- Case Functions & TEXTJOIN").font = Font(
        bold=True, size=11, color="2F5496")

    bonus_headers = ["Input", "UPPER", "LOWER", "PROPER"]
    for i, h in enumerate(bonus_headers, 1):
        ws.cell(row=bonus_row + 1, column=i, value=h)
    style_header_row(ws, bonus_row + 1, 4)

    bonus_inputs = ["hello world", "ALREADY UPPER", "mIxEd CaSe"]
    for i, val in enumerate(bonus_inputs):
        r = bonus_row + 2 + i
        ws.cell(row=r, column=1, value=val)
        for c, func in enumerate(["UPPER", "LOWER", "PROPER"], 2):
            cell_ref = f"A{r}"
            fc = ws.cell(row=r, column=c, value=f"={func}({cell_ref})")
            mark_formula(fc, f"{func} converts text case")

    auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 4 -- Date Functions
# ---------------------------------------------------------------------------

def build_date_sheet(wb):
    ws = wb.create_sheet("Date Functions")

    write_title(ws, 1, "Date & Time Function Examples",
                "NETWORKDAYS, EOMONTH, DATEDIF, WORKDAY with project timeline calculations", 8)

    # --- Project timeline data ---
    data_row = 4
    headers = ["Project", "Start Date", "End Date", "Status"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=data_row, column=i, value=h)
    style_header_row(ws, data_row, 4)

    projects = [
        ["Website Redesign",   date(2025, 1, 6),  date(2025, 3, 28), "Complete"],
        ["Mobile App MVP",     date(2025, 2, 10), date(2025, 6, 30), "In Progress"],
        ["Data Migration",     date(2025, 4, 1),  date(2025, 5, 15), "In Progress"],
        ["API Integration",    date(2025, 3, 3),  date(2025, 4, 18), "Complete"],
        ["Dashboard Build",    date(2025, 5, 1),  date(2025, 8, 29), "Planned"],
    ]
    for r, row_data in enumerate(projects, data_row + 1):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in (2, 3):
                cell.number_format = "YYYY-MM-DD"
    apply_data_borders(ws, data_row, data_row + len(projects), 1, 4)

    # --- Formula demo area (columns F-H) ---
    f_row = data_row
    ws.cell(row=f_row, column=6, value="Calculation")
    ws.cell(row=f_row, column=7, value="Formula")
    ws.cell(row=f_row, column=8, value="Result")
    for c in range(6, 9):
        ws.cell(row=f_row, column=c).font = SUBHEADER_FONT
        ws.cell(row=f_row, column=c).fill = SUBHEADER_FILL
        ws.cell(row=f_row, column=c).alignment = Alignment(horizontal="center")
        ws.cell(row=f_row, column=c).border = THIN_BORDER

    date_formulas = [
        ("Working days (Project 1)",
         "NETWORKDAYS",
         "=NETWORKDAYS(B5,C5)",
         "NETWORKDAYS counts business days between two dates (excludes weekends)"),
        ("End of month from start (Proj 2)",
         "EOMONTH",
         "=EOMONTH(B6,0)",
         "EOMONTH returns the last day of the month; 0 = same month, 1 = next month, etc."),
        ("Months between dates (Proj 1)",
         "DATEDIF months",
         '=DATEDIF(B5,C5,"M")',
         'DATEDIF(start,end,"M") returns complete months between two dates'),
        ("Days between dates (Proj 1)",
         "DATEDIF days",
         '=DATEDIF(B5,C5,"D")',
         'DATEDIF with "D" returns total calendar days'),
        ("30 workdays after Proj 3 start",
         "WORKDAY",
         "=WORKDAY(B7,30)",
         "WORKDAY adds N business days to a start date (skipping weekends)"),
        ("Next month-end from Proj 4 start",
         "EOMONTH +1",
         "=EOMONTH(B8,1)",
         "EOMONTH with months=1 jumps to the end of the following month"),
        ("Today's date",
         "TODAY",
         "=TODAY()",
         "TODAY() returns the current date -- updates every time the workbook recalculates"),
        ("Days until Proj 5 ends",
         "Days remaining",
         "=C9-TODAY()",
         "Simple subtraction gives calendar days remaining from today"),
    ]

    for i, (desc, label, formula, comment) in enumerate(date_formulas):
        r = f_row + 1 + i
        ws.cell(row=r, column=6, value=desc)
        ws.cell(row=r, column=7, value=label)
        result_cell = ws.cell(row=r, column=8, value=formula)
        result_cell.number_format = "YYYY-MM-DD"
        mark_formula(result_cell, comment)

    # --- Bonus: Quarter calculation ---
    bonus_row = f_row + len(date_formulas) + 3
    ws.merge_cells(start_row=bonus_row, start_column=1, end_row=bonus_row, end_column=4)
    ws.cell(row=bonus_row, column=1,
            value="Bonus -- Fiscal Quarter from Date").font = Font(
        bold=True, size=11, color="2F5496")

    ws.cell(row=bonus_row + 1, column=1, value="Date")
    ws.cell(row=bonus_row + 1, column=2, value="Quarter")
    style_header_row(ws, bonus_row + 1, 2)

    sample_dates = [date(2025, 1, 15), date(2025, 5, 20), date(2025, 9, 1), date(2025, 12, 31)]
    for i, d in enumerate(sample_dates):
        r = bonus_row + 2 + i
        ws.cell(row=r, column=1, value=d).number_format = "YYYY-MM-DD"
        qcell = ws.cell(row=r, column=2,
                        value=f'="Q"&INT((MONTH(A{r})-1)/3)+1')
        mark_formula(qcell, "Calculate fiscal quarter: Q1-Q4 from the month number")

    auto_width(ws)


# ---------------------------------------------------------------------------
# Main -- assemble workbook
# ---------------------------------------------------------------------------

def main():
    wb = Workbook()
    # Remove the default blank sheet created by openpyxl
    wb.remove(wb.active)

    build_lookup_sheet(wb)
    build_conditional_sheet(wb)
    build_text_sheet(wb)
    build_date_sheet(wb)

    output_path = "formula_toolkit.xlsx"
    wb.save(output_path)
    print(f"[OK] Generated '{output_path}' with {len(wb.sheetnames)} sheets:")
    for name in wb.sheetnames:
        print(f"     - {name}")


if __name__ == "__main__":
    main()
