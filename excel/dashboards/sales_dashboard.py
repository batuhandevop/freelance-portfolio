"""
Sales Dashboard Generator
=========================
Generates a professional Excel sales dashboard (sales_dashboard.xlsx) with:
  - A "Data" sheet containing 120 rows of realistic 2024 sales transactions
  - A "Dashboard" sheet with KPI metrics, regional/category/monthly summaries,
    conditional formatting, and polished styling

Requirements:
    pip install openpyxl

Usage:
    python sales_dashboard.py
"""

import random
import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

REGIONS = ["North", "South", "East", "West"]

PRODUCTS = {
    # product_name: (category, base_price)
    "Laptop":      ("Electronics",  899.00),
    "Monitor":     ("Electronics",  349.00),
    "Keyboard":    ("Peripherals",   79.00),
    "Mouse":       ("Peripherals",   49.00),
    "Headset":     ("Accessories",  129.00),
    "Webcam":      ("Accessories",   89.00),
    "USB Hub":     ("Peripherals",   39.00),
    "Desk Lamp":   ("Office",        59.00),
    "Chair":       ("Office",       299.00),
    "Mousepad":    ("Accessories",   19.00),
}

SALESPERSONS = [
    "Alice Johnson", "Bob Martinez", "Carol Lee",
    "David Kim", "Emma Wilson", "Frank Chen",
]

NUM_ROWS = 120
SEED = 42

# Styling constants
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(bold=True, color="1F4E79", size=12)
TITLE_FONT = Font(bold=True, color="1F4E79", size=16)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CURRENCY_FMT = '#,##0.00 "$"'
NUMBER_FMT = "#,##0"

# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------


def generate_sales_data(rng: random.Random):
    """Return a list of dicts representing individual sales transactions."""
    rows = []
    product_names = list(PRODUCTS.keys())
    start_date = datetime.date(2024, 1, 1)
    end_date = datetime.date(2024, 12, 31)
    days_range = (end_date - start_date).days

    for _ in range(NUM_ROWS):
        product = rng.choice(product_names)
        category, base_price = PRODUCTS[product]
        # Add +/-15 % price variation to keep it realistic
        unit_price = round(base_price * rng.uniform(0.85, 1.15), 2)
        quantity = rng.randint(1, 20)
        revenue = round(unit_price * quantity, 2)
        order_date = start_date + datetime.timedelta(days=rng.randint(0, days_range))

        rows.append({
            "Date": order_date,
            "Region": rng.choice(REGIONS),
            "Product": product,
            "Category": category,
            "Quantity": quantity,
            "Unit Price": unit_price,
            "Revenue": revenue,
            "Salesperson": rng.choice(SALESPERSONS),
        })

    # Sort by date for readability
    rows.sort(key=lambda r: r["Date"])
    return rows


def style_header_row(ws, row, num_cols):
    """Apply header styling to a row of cells."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def apply_border_range(ws, min_row, max_row, min_col, max_col):
    """Apply thin borders to a rectangular range."""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def write_section_header(ws, row, col, text, width=2):
    """Write a colored section header spanning `width` columns."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="left")
    for c in range(col, col + width):
        ws.cell(row=row, column=c).fill = SECTION_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER


# ---------------------------------------------------------------------------
# Build the Data sheet
# ---------------------------------------------------------------------------


def build_data_sheet(wb, data):
    """Populate the 'Data' sheet with raw transaction data."""
    ws = wb.active
    ws.title = "Data"

    headers = ["Date", "Region", "Product", "Category",
               "Quantity", "Unit Price", "Revenue", "Salesperson"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for row_dict in data:
        ws.append([row_dict[h] for h in headers])

    num_rows = len(data) + 1  # including header

    # Column widths
    col_widths = [12, 10, 14, 14, 10, 12, 14, 18]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Format dates, currency, numbers
    for r in range(2, num_rows + 1):
        ws.cell(row=r, column=1).number_format = "YYYY-MM-DD"
        ws.cell(row=r, column=5).number_format = NUMBER_FMT
        ws.cell(row=r, column=6).number_format = CURRENCY_FMT
        ws.cell(row=r, column=7).number_format = CURRENCY_FMT
        # Light borders
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = THIN_BORDER

    # Freeze top row
    ws.freeze_panes = "A2"

    return ws


# ---------------------------------------------------------------------------
# Build the Dashboard sheet
# ---------------------------------------------------------------------------


def build_dashboard_sheet(wb, num_data_rows):
    """Create the 'Dashboard' sheet with formulas and formatting."""
    ws = wb.create_sheet("Dashboard")
    last_data_row = num_data_rows + 1  # +1 for header

    # Reference helpers (pointing to the Data sheet)
    rev_range = f"Data!G2:G{last_data_row}"
    qty_range = f"Data!E2:E{last_data_row}"
    region_range = f"Data!B2:B{last_data_row}"
    category_range = f"Data!D2:D{last_data_row}"
    date_range = f"Data!A2:A{last_data_row}"

    current_row = 1

    # ---- Title ----
    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="Sales Dashboard - 2024")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")
    current_row = 3

    # ---- KPI Section ----
    write_section_header(ws, current_row, 1, "Key Performance Indicators", width=4)
    current_row += 1

    kpis = [
        ("Total Revenue",       f'=SUM({rev_range})'),
        ("Total Orders",        f'=COUNTA({rev_range})'),
        ("Total Units Sold",    f'=SUM({qty_range})'),
        ("Avg Order Value",     f'=AVERAGE({rev_range})'),
    ]
    kpi_formats = [CURRENCY_FMT, NUMBER_FMT, NUMBER_FMT, CURRENCY_FMT]

    for i, (label, formula) in enumerate(kpis):
        col_start = 1 + i * 2
        lbl_cell = ws.cell(row=current_row, column=col_start, value=label)
        lbl_cell.font = Font(bold=True, size=10, color="1F4E79")
        lbl_cell.alignment = Alignment(horizontal="center")
        lbl_cell.border = THIN_BORDER

        val_cell = ws.cell(row=current_row + 1, column=col_start, value=formula)
        val_cell.font = Font(bold=True, size=14, color="1F4E79")
        val_cell.number_format = kpi_formats[i]
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.border = THIN_BORDER

    current_row += 3

    # ---- Revenue by Region ----
    write_section_header(ws, current_row, 1, "Revenue by Region", width=3)
    current_row += 1

    # Sub-headers
    for ci, hdr in enumerate(["Region", "Revenue", "% of Total"], start=1):
        c = ws.cell(row=current_row, column=ci, value=hdr)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
    current_row += 1

    region_start_row = current_row
    for region in REGIONS:
        ws.cell(row=current_row, column=1, value=region).border = THIN_BORDER

        rev_cell = ws.cell(
            row=current_row, column=2,
            value=f'=SUMIFS({rev_range},{region_range},A{current_row})'
        )
        rev_cell.number_format = CURRENCY_FMT
        rev_cell.border = THIN_BORDER

        pct_cell = ws.cell(
            row=current_row, column=3,
            value=f'=IF(SUM({rev_range})=0,0,B{current_row}/SUM({rev_range}))'
        )
        pct_cell.number_format = "0.0%"
        pct_cell.alignment = Alignment(horizontal="center")
        pct_cell.border = THIN_BORDER

        current_row += 1
    region_end_row = current_row - 1

    current_row += 1

    # ---- Revenue by Product Category ----
    write_section_header(ws, current_row, 1, "Revenue by Product Category", width=3)
    current_row += 1

    for ci, hdr in enumerate(["Category", "Revenue", "% of Total"], start=1):
        c = ws.cell(row=current_row, column=ci, value=hdr)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
    current_row += 1

    categories = sorted(set(v[0] for v in PRODUCTS.values()))
    cat_start_row = current_row
    for cat in categories:
        ws.cell(row=current_row, column=1, value=cat).border = THIN_BORDER

        rev_cell = ws.cell(
            row=current_row, column=2,
            value=f'=SUMIFS({rev_range},{category_range},A{current_row})'
        )
        rev_cell.number_format = CURRENCY_FMT
        rev_cell.border = THIN_BORDER

        pct_cell = ws.cell(
            row=current_row, column=3,
            value=f'=IF(SUM({rev_range})=0,0,B{current_row}/SUM({rev_range}))'
        )
        pct_cell.number_format = "0.0%"
        pct_cell.alignment = Alignment(horizontal="center")
        pct_cell.border = THIN_BORDER

        current_row += 1
    cat_end_row = current_row - 1

    current_row += 1

    # ---- Monthly Revenue ----
    write_section_header(ws, current_row, 1, "Monthly Revenue Trend", width=3)
    current_row += 1

    for ci, hdr in enumerate(["Month", "Revenue", "Orders"], start=1):
        c = ws.cell(row=current_row, column=ci, value=hdr)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER
    current_row += 1

    month_names = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    month_start_row = current_row
    for m_idx, m_name in enumerate(month_names, start=1):
        m_start = f"DATE(2024,{m_idx},1)"
        # End date: first day of next month minus 1, or Dec 31 for December
        if m_idx < 12:
            m_end = f"DATE(2024,{m_idx + 1},1)-1"
        else:
            m_end = "DATE(2024,12,31)"

        ws.cell(row=current_row, column=1, value=m_name).border = THIN_BORDER

        rev_cell = ws.cell(
            row=current_row, column=2,
            value=f'=SUMIFS({rev_range},{date_range},">="&{m_start},{date_range},"<="&{m_end})'
        )
        rev_cell.number_format = CURRENCY_FMT
        rev_cell.border = THIN_BORDER

        ord_cell = ws.cell(
            row=current_row, column=3,
            value=f'=COUNTIFS({date_range},">="&{m_start},{date_range},"<="&{m_end})'
        )
        ord_cell.number_format = NUMBER_FMT
        ord_cell.alignment = Alignment(horizontal="center")
        ord_cell.border = THIN_BORDER

        current_row += 1
    month_end_row = current_row - 1

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    # ---- Conditional Formatting ----

    # Color scale on regional revenue (green = high, red = low)
    ws.conditional_formatting.add(
        f"B{region_start_row}:B{region_end_row}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        )
    )

    # Color scale on category revenue
    ws.conditional_formatting.add(
        f"B{cat_start_row}:B{cat_end_row}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        )
    )

    # Data bars on monthly revenue
    ws.conditional_formatting.add(
        f"B{month_start_row}:B{month_end_row}",
        DataBarRule(
            start_type="min", end_type="max",
            color="5B9BD5", showValue=True,
        )
    )

    # Data bars on monthly orders
    ws.conditional_formatting.add(
        f"C{month_start_row}:C{month_end_row}",
        DataBarRule(
            start_type="min", end_type="max",
            color="ED7D31", showValue=True,
        )
    )

    # Freeze pane below the title
    ws.freeze_panes = "A3"

    # Set Dashboard as the active sheet when the file opens
    wb.active = wb.sheetnames.index("Dashboard")

    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    rng = random.Random(SEED)
    data = generate_sales_data(rng)

    wb = Workbook()
    build_data_sheet(wb, data)
    build_dashboard_sheet(wb, len(data))

    filename = "sales_dashboard.xlsx"
    wb.save(filename)
    print(f"Dashboard saved to {filename}")
    print(f"  - {len(data)} rows of sales data on the 'Data' sheet")
    print(f"  - Summary dashboard with formulas on the 'Dashboard' sheet")
    print(f"Open the file in Excel or LibreOffice Calc to see the results.")


if __name__ == "__main__":
    main()
